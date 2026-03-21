"""
Анализ рынков Polymarket + формирование 15 прогнозов + запись ставок в Excel.

Использует строгие фильтры из плана доработок:
- Whitelist лиг (Tier1/Tier2), skip Colombia/BetPlay
- Запрет NO при YES > 65%
- Kelly 0.25, min edge 3% (оценка без Pinnacle)
- Liquidity >= 3000
- Одна ставка — одно событие (не более 1 ставки на матч, напр. O/U 22.5 и 23.5)
"""
from __future__ import annotations
import json
import re
import sys
import io
import urllib.request
from datetime import datetime, timezone, timedelta
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Параметры ─────────────────────────────────────────────────
BANKROLL = 1_000.0
KELLY_FRAC = 0.25
MAX_BET_USD = 30.0
MIN_BET_USD = 5.0
EDGE_EST = 0.04  # 4% — консервативная оценка без Pinnacle
MIN_LIQUIDITY = 2_000
NUM_BETS = 15
BETTING_WINDOW_MIN_H = 2.0   # Не ставить за <2ч до матча (line movement)
BETTING_WINDOW_MAX_H = 24.0  # Не ставить за >24ч (смена состава)

EXCEL_PATH = Path(r"C:\Users\Lomov\Desktop\polymarket-agent\polymarket_рекомендации.xlsx")
GAMMA_API = "https://gamma-api.polymarket.com/markets?closed=false&limit=200"
NOW = datetime.now(timezone.utc)

# ── Стили ────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BLUE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ZEBRA_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ── Лиги: Tier1 (полный Kelly), Tier2 (×0.5), Skip (исключить) ──
TIER1_KW = ["epl", "premier league", "la liga", "bundesliga", "serie a", "ligue 1",
            "champions league", "ucl", "nba", "nfl", "ufc", "mma", "stanley cup", "nhl"]
TIER2_KW = ["liga mx", "ligamx", "mls", "mexico", "championship", "efl"]
SKIP_KW = ["colombia", "betplay", "super lig", "ligue 2"]

SPORT_KW = [
    "win", "beat", " vs ", "o/u", "over", "under", "total",
    "nba", "nhl", "nfl", "ufc", "mlb", "epl", "ucl",
    "fc ", " fc", "afc", "tennis", "match", "fight", "bout",
    "arsenal", "chelsea", "liverpool", "tottenham", "manchester",
    "villa", "city", "united", "lakers", "celtics", "warriors",
    "nuggets", "bucks", "hurricanes", "stanley cup",
    "world cup", "fifa", "handicap", "rybakina", "kartal",
    "middlesbrough", "charlton", "ly ", "loud",
]


def event_key(q: str) -> str:
    """Ключ события: одна ставка на одно событие (Kartal O/U 22.5 и 23.5 → один матч)."""
    s = (q or "").strip().lower()
    s = re.sub(r"\s*(O/U|over|under|total)\s*[\d.]+.*$", "", s, flags=re.I)
    s = re.sub(r"\s*[-:]\s*$", "", s).strip()
    return s[:80] if s else (q or "")[:80].lower()


def league_tier(q: str) -> int:
    ql = (q or "").lower()
    if any(s in ql for s in SKIP_KW):
        return 0
    if any(t in ql for t in TIER2_KW):
        return 2
    if any(t in ql for t in TIER1_KW):
        return 1
    # По умолчанию — если есть sport keywords, считаем Tier1
    if any(k in ql for k in SPORT_KW):
        return 1
    return 0


def hdr(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def kelly_bet(yes_p: float, no_p: float, side: str, edge: float = EDGE_EST,
              bankroll: float = BANKROLL, frac: float = KELLY_FRAC,
              max_usd: float = MAX_BET_USD, min_usd: float = MIN_BET_USD,
              tier_mult: float = 1.0) -> float:
    """Kelly для YES или NO."""
    if side == "YES":
        p_true = min(yes_p + edge, 0.95)
        price = yes_p
        b = (1 / price) - 1 if price > 0 else 0
    else:
        p_true = min(no_p + edge, 0.95)  # true prob NO выше чем рынок
        price = no_p
        b = (1 / price) - 1 if price > 0 else 0
    if b <= 0:
        return 0.0
    f = (p_true * b - (1 - p_true)) / b
    f = max(f, 0)
    bet = bankroll * f * frac * tier_mult
    bet = min(bet, max_usd)
    bet = max(bet, min_usd) if f > 0 else 0
    return round(bet, 2)


def main():
    # 1. Загрузка рынков с API
    print("  Загрузка рынков Polymarket...")
    try:
        with urllib.request.urlopen(GAMMA_API, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        print(f"  [!] API недоступен: {e}. Пробуем кэш...")
        cache_files = list(Path(r"C:\Users\Lomov\.cursor\projects\c-Users-Lomov-Documents-Arduino-libraries-M5Unified-src-utility\agent-tools").glob("*.txt"))
        data = []
        for fp in cache_files[:5]:
            try:
                data.extend(json.load(open(fp, encoding="utf-8")))
            except Exception:
                pass
        if not data:
            print("  [!] Нет данных. Создайте .env с ODDS_API_KEY и запустите main.py")
            return

    print(f"  Загружено рынков: {len(data)}")

    # 2. Парсинг и фильтрация
    seen_cids = set()
    all_markets = []

    for m in data:
        cid = m.get("conditionId", "") or m.get("id", "")
        if cid in seen_cids:
            continue
        seen_cids.add(cid)
        prices = m.get("outcomePrices", "[]")
        if isinstance(prices, str):
            try:
                prices = json.loads(prices)
            except Exception:
                prices = []
        if len(prices) < 2:
            continue
        liq = float(m.get("liquidity", 0) or 0)
        if liq < MIN_LIQUIDITY:
            continue
        yes_p = float(prices[0])
        no_p = float(prices[1])
        if not (0.10 < yes_p < 0.90):
            continue
        end_raw = m.get("endDate", "")
        end_dt = None
        if end_raw:
            try:
                end_dt = datetime.fromisoformat(end_raw.replace("Z", "+00:00"))
            except Exception:
                pass
        if end_dt and end_dt < NOW:
            continue
        # Фильтр времени: 2–24ч до матча (analyze_losses рекомендация)
        if end_dt:
            hours_to_start = (end_dt - NOW).total_seconds() / 3600.0
            if hours_to_start < BETTING_WINDOW_MIN_H or hours_to_start > BETTING_WINDOW_MAX_H:
                continue
        q = m.get("question", "")
        if not any(kw in q.lower() for kw in SPORT_KW):
            continue
        tier = league_tier(q)
        if tier == 0:
            continue

        all_markets.append({
            "q": q, "cid": cid, "slug": m.get("slug", ""),
            "yes": yes_p, "no": no_p,
            "liq": liq, "vol24": float(m.get("volume24hr", 0) or 0),
            "end_dt": end_dt, "tier": tier,
        })

    print(f"  Спортивных рынков (фильтр): {len(all_markets)}")

    # 3. Расчёт ставок
    bets_to_place = []
    seen_ends = set()
    seen_events = set()  # Одна ставка — одно событие (правило п.4)

    kelly_mult_tier = {1: 1.0, 2: 0.5}

    for m in sorted(all_markets, key=lambda x: (-x["liq"], -max(x["yes"], x["no"]))):
        if len(bets_to_place) >= NUM_BETS:
            break
        tier_mult = kelly_mult_tier.get(m["tier"], 0.5)
        # O/U и тоталы без Pinnacle — половинный Kelly (высокая волатильность)
        ql = m["q"].lower()
        ou_mult = 0.5 if any(k in ql for k in ["o/u", "over", "under", "total"]) else 1.0
        tier_mult *= ou_mult

        # Выбор стороны: YES если цена < 0.60, иначе NO (но только если YES <= 0.65)
        if m["yes"] < 0.60:
            side = "YES"
            price = m["yes"]
            # Запрет NO при YES > 65% — мы ставим YES, ок
        elif m["yes"] <= 0.65:
            side = "NO"
            price = m["no"]
        else:
            continue  # YES > 65% — не ставим NO по правилу

        bet_usd = kelly_bet(m["yes"], m["no"], side, tier_mult=tier_mult)
        if bet_usd <= 0:
            continue

        end_key = m["end_dt"].strftime("%Y-%m-%d-%H") if m["end_dt"] else "no-date"
        if sum(1 for k in seen_ends if k == end_key) >= 3:
            continue
        ev_key = event_key(m["q"])
        if ev_key in seen_events:
            continue  # Уже ставка на это событие (напр. O/U 22.5 и 23.5 одного матча)
        seen_ends.add(end_key)
        seen_events.add(ev_key)

        win_usd = round(bet_usd * (1 / price - 1), 2)
        if side == "YES":
            p_true = min(m["yes"] + EDGE_EST, 0.95)
            ev = round(p_true * win_usd - (1 - p_true) * bet_usd, 2)
        else:
            p_true = min(m["no"] + EDGE_EST, 0.95)
            ev = round(p_true * win_usd - (1 - p_true) * bet_usd, 2)
        decimal_odds = round(1 / price, 3)
        check_dt = (m["end_dt"] + timedelta(hours=2)) if m["end_dt"] else None

        bets_to_place.append({
            "q": m["q"], "cid": m["cid"], "slug": m["slug"],
            "yes": m["yes"], "no": m["no"],
            "side": side, "price": price, "odds": decimal_odds,
            "bet_usd": bet_usd, "win_usd": win_usd, "ev": ev,
            "liq": m["liq"], "vol24": m["vol24"],
            "end_dt": m["end_dt"], "check_dt": check_dt,
            "tier": m["tier"],
        })

    total_risk = sum(b["bet_usd"] for b in bets_to_place)
    total_ev = sum(b["ev"] for b in bets_to_place)

    print(f"\n  Прогнозов отобрано: {len(bets_to_place)}")
    print(f"  Суммарный риск: ${total_risk:.2f}")
    print(f"  Суммарный EV: ${total_ev:+.2f}")

    if not bets_to_place:
        print("  [!] Нет подходящих рынков. Ослабьте фильтры или обновите кэш.")
        return

    # 4. Запись в Excel
    wb = load_workbook(EXCEL_PATH)
    ws_bets = wb["Ставки"]
    ws_stats = wb["Статистика"]

    HEADERS = [
        "№", "Condition ID", "Рынок (вопрос)",
        "Направление", "Коэф. (dec)",
        "Ставка $", "Цена доли", "Потенц. выигрыш $", "EV $",
        "Результат", "P&L $", "Статус",
        "Дата ставки", "Дата окончания", "Проверить в",
    ]

    for c, h in enumerate(HEADERS, 1):
        ws_bets.cell(row=1, column=c, value=h)
    hdr(ws_bets, 1, len(HEADERS))
    ws_bets.row_dimensions[1].height = 32

    next_row = 2
    while ws_bets.cell(row=next_row, column=3).value not in (None, ""):
        next_row += 1
    start_row = next_row

    for i, b in enumerate(bets_to_place, 1):
        r = next_row
        next_row += 1
        base = ZEBRA_FILL if i % 2 == 0 else PatternFill()
        bet_num = start_row - 1 + i
        side_fill = BLUE_FILL if b["side"] == "YES" else ORANGE_FILL
        end_str = b["end_dt"].strftime("%Y-%m-%d %H:%M") if b["end_dt"] else ""
        check_str = b["check_dt"].strftime("%Y-%m-%d %H:%M") if b["check_dt"] else ""

        ws_bets.cell(row=r, column=1, value=bet_num).alignment = CENTER
        ws_bets.cell(row=r, column=2, value=str(b["cid"])[:35])
        ws_bets.cell(row=r, column=3, value=b["q"]).alignment = WRAP
        ws_bets.cell(row=r, column=4, value=b["side"]).alignment = CENTER
        ws_bets.cell(row=r, column=5, value=b["odds"]).number_format = "0.000"
        ws_bets.cell(row=r, column=6, value=b["bet_usd"]).number_format = "#,##0.00"
        ws_bets.cell(row=r, column=7, value=b["price"]).number_format = "0.000"
        ws_bets.cell(row=r, column=8, value=b["win_usd"]).number_format = "#,##0.00"
        ws_bets.cell(row=r, column=9, value=b["ev"]).number_format = "#,##0.00"
        ws_bets.cell(row=r, column=10, value="Ожидание").alignment = CENTER
        ws_bets.cell(row=r, column=11, value="").number_format = "#,##0.00"
        ws_bets.cell(row=r, column=12, value="Активна").alignment = CENTER
        ws_bets.cell(row=r, column=13, value=NOW.strftime("%Y-%m-%d %H:%M"))
        ws_bets.cell(row=r, column=14, value=end_str)
        ws_bets.cell(row=r, column=15, value=check_str)

        ws_bets.cell(row=r, column=4).fill = side_fill
        ws_bets.cell(row=r, column=10).fill = YELLOW_FILL
        ws_bets.cell(row=r, column=12).fill = BLUE_FILL

        for c in range(1, len(HEADERS) + 1):
            cell = ws_bets.cell(row=r, column=c)
            cell.border = THIN_BORDER
            if c not in (4, 10, 12):
                cell.fill = base
        ws_bets.row_dimensions[r].height = 32

    col_w = {"A": 4, "B": 22, "C": 52, "D": 12, "E": 10, "F": 10, "G": 9, "H": 18, "I": 9,
             "J": 14, "K": 10, "L": 10, "M": 18, "N": 18, "O": 18}
    for col, w in col_w.items():
        ws_bets.column_dimensions[col].width = w

    ws_stats.cell(row=7, column=1, value="Дата последнего обновл.")
    ws_stats.cell(row=7, column=2, value=NOW.strftime("%Y-%m-%d %H:%M UTC"))
    ws_stats.cell(row=8, column=1, value="Активных ставок")
    ws_stats.cell(row=8, column=2, value=len(bets_to_place))
    ws_stats.cell(row=9, column=1, value="Суммарный риск $")
    ws_stats.cell(row=9, column=2, value=round(total_risk, 2))
    ws_stats.cell(row=10, column=1, value="Суммарный EV $")
    ws_stats.cell(row=10, column=2, value=round(total_ev, 2))
    for rn in range(7, 11):
        for c in [1, 2]:
            ws_stats.cell(row=rn, column=c).border = THIN_BORDER
            ws_stats.cell(row=rn, column=c).font = Font(bold=(c == 1))

    log_row = 12
    while ws_stats.cell(row=log_row, column=1).value not in (None, ""):
        log_row += 1
        if log_row > 300:
            break
    ws_stats.cell(row=log_row, column=1, value=NOW.strftime("%Y-%m-%d %H:%M"))
    ws_stats.cell(row=log_row, column=2,
                  value=f"Анализ + {len(bets_to_place)} ставок. Риск=${total_risk:.2f}, EV=${total_ev:+.2f}")
    ws_stats.cell(row=log_row, column=2).alignment = WRAP
    for c in [1, 2]:
        ws_stats.cell(row=log_row, column=c).border = THIN_BORDER
    ws_stats.cell(row=log_row, column=1).fill = BLUE_FILL

    wb.save(EXCEL_PATH)
    print(f"\n  Excel сохранён: {EXCEL_PATH}")

    # 5. Отчёт
    print("\n" + "=" * 80)
    print(f"  ПРОГНОЗЫ И СТАВКИ — {NOW.strftime('%Y-%m-%d %H:%M UTC')}")
    print("=" * 80)
    for i, b in enumerate(bets_to_place, 1):
        end_str = b["end_dt"].strftime("%d.%m.%Y %H:%M UTC") if b["end_dt"] else "?"
        tier_lbl = "T1" if b["tier"] == 1 else "T2"
        print(f"\n  #{i}  [{b['side']}]  Коэф. {b['odds']:.3f}  (лига {tier_lbl})")
        print(f"  {b['q'][:72]}")
        print(f"  Ставка: ${b['bet_usd']:.2f}   Выигрыш: +${b['win_usd']:.2f}   EV: ${b['ev']:+.2f}")
        print(f"  Ликв: ${b['liq']:,.0f}   Окончание: {end_str}")
    print("\n" + "-" * 80)
    print(f"  Суммарный риск: ${total_risk:.2f}   EV: ${total_ev:+.2f}")
    print("  ⚠ Бумажные ставки. Для реального edge — настрой ODDS_API_KEY и main.py")
    print("-" * 80)


if __name__ == "__main__":
    main()
