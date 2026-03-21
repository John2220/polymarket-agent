"""
Анализ отрицательного результата — 10 причин + план улучшений.
"""
from __future__ import annotations
import sys, io
from datetime import datetime, timezone
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH = Path(r"C:\Users\Lomov\Desktop\polymarket-agent\polymarket_рекомендации.xlsx")
NOW = datetime.now(timezone.utc)

# ── Стили ────────────────────────────────────────────────────
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=12)
HEADER_FILL   = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
RED_FILL      = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_DARK      = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
GREEN_FILL    = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_DARK    = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
YELLOW_FILL   = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL   = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
BLUE_FILL     = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
BLUE_DARK     = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
GRAY_FILL     = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
PURPLE_FILL   = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP          = Alignment(wrap_text=True, vertical="top")
THIN          = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
MEDIUM        = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"),  bottom=Side(style="medium"),
)

def cell(ws, r, c, val="", bold=False, fill=None, align=None, fmt=None, size=11,
         color="000000", border=THIN):
    cl = ws.cell(row=r, column=c, value=val)
    cl.font      = Font(bold=bold, size=size, color=color)
    cl.alignment = align or WRAP
    cl.border    = border
    if fill: cl.fill = fill
    if fmt:  cl.number_format = fmt
    return cl

# ════════════════════════════════════════════════════════════
# ДАННЫЕ ПО СТАВКАМ
# ════════════════════════════════════════════════════════════

bets = [
    {
        "num": 1,
        "q":       "Aston Villa FC win (27.02.2026)",
        "side":    "YES",
        "price":   0.48,
        "bet":     10.00,
        "pnl":    -10.00,
        "outcome": "Ничья 1:1",
        "won":     False,
        "errors":  ["Нет источника edge — поставили по рыночной цене",
                    "Ничья не учтена в бинарном рынке YES/NO",
                    "Цена 0.48 — слишком близко к 50%, минимальный edge"],
    },
    {
        "num": 2,
        "q":       "CD Tolima win (04.03.2026)",
        "side":    "YES",
        "price":   0.295,
        "bet":     17.73,
        "pnl":    -17.73,
        "outcome": "Ничья 1:1",
        "won":     False,
        "errors":  ["Андердог без анализа формы и состава команды",
                    "Ничья = проигрыш для ставки YES на победу",
                    "Нет данных от Pinnacle для сравнения коэффициентов"],
    },
    {
        "num": 3,
        "q":       "Puebla vs Tigres O/U 1.5 (NO=UNDER)",
        "side":    "NO",
        "price":   0.245,
        "bet":     30.00,
        "pnl":    -30.00,
        "outcome": "0:2 Tigres (2 гола — OVER сыграл)",
        "won":     False,
        "errors":  ["Рынок давал OVER 75.5% — мы ставили против рынка без обоснования",
                    "O/U 1.5 — очень низкий барьер, 75%+ матчей идут OVER",
                    "Максимальная ставка $30 на самый рискованный сигнал"],
    },
    {
        "num": 4,
        "q":       "Club Tijuana win (04.03.2026)",
        "side":    "YES",
        "price":   0.265,
        "bet":     17.01,
        "pnl":    +47.18,
        "outcome": "2:0 Querétaro — победа",
        "won":     True,
        "errors":  [],
    },
]

closed = [b for b in bets]
won_bets  = [b for b in closed if b["won"]]
lost_bets = [b for b in closed if not b["won"]]
total_risk = sum(b["bet"] for b in closed)
total_pnl  = sum(b["pnl"] for b in closed)
roi = total_pnl / total_risk * 100

print("=" * 70)
print("  АНАЛИЗ РЕЗУЛЬТАТОВ")
print("=" * 70)
print(f"  Ставок: {len(closed)}  Побед: {len(won_bets)}  Проигрышей: {len(lost_bets)}")
print(f"  WR: {len(won_bets)/len(closed)*100:.1f}%  |  P&L: ${total_pnl:.2f}  |  ROI: {roi:.1f}%")
print()

# ════════════════════════════════════════════════════════════
# 10 ПРИЧИН ОТРИЦАТЕЛЬНОГО РЕЗУЛЬТАТА
# ════════════════════════════════════════════════════════════

REASONS = [
    {
        "n":      1,
        "title":  "Нет реального источника edge (Pinnacle / sharp линии)",
        "detail": (
            "Мы оценивали edge в +5% «вручную», без сравнения с Pinnacle или другим "
            "острым букмекером. В итоге мы ставили по рыночной цене Polymarket, "
            "не зная истинной вероятности. Edge = 0 → математическое ожидание отрицательное "
            "из-за slippage и комиссий."
        ),
        "fix": (
            "Подключить The Odds API (ключ уже в .env). Функция odds_api.py и match_odds_to_markets() "
            "уже написаны, но не используются. Включить сравнение: если Polymarket YES < Pinnacle "
            "implied probability — есть реальный edge. Ставить ТОЛЬКО при edge > 3%."
        ),
        "priority": "КРИТИЧНО",
        "fill": RED_FILL,
    },
    {
        "n":      2,
        "title":  "Ставки на андердогов без фундаментального анализа",
        "detail": (
            "Tolima (0.295), Tijuana (0.265), Aston Villa (0.48) — все три ставки на команды, "
            "которые рынок оценивает ниже 50%. Мы выбрали их только по высокому коэффициенту "
            "(потенциал выигрыша), не анализируя форму, состав, домашний/гостевой фактор, H2H."
        ),
        "fix": (
            "Добавить модуль data/team_stats.py: парсинг последних 5 матчей из FlashScore API / "
            "Football-data.org API (бесплатно). Смотреть: форма (W/D/L), xG, голы дома/в гостях. "
            "Не ставить на команду с формой хуже 40% в последних 5 играх."
        ),
        "priority": "КРИТИЧНО",
        "fill": RED_FILL,
    },
    {
        "n":      3,
        "title":  "Ничья не отражена в стратегии YES/NO",
        "detail": (
            "2 из 3 проигрышей (Aston Villa, Tolima) — ничьи. В бинарном рынке Polymarket "
            "'Will X win?' ответ YES = только победа. Ничья всегда = NO. "
            "Базовая частота ничьих в европейском футболе: 25-30%. В Liga BetPlay ещё выше. "
            "Мы не учли это в расчёте истинной вероятности."
        ),
        "fix": (
            "В kelly.py добавить 3-way correction: p_win = p_yes / (1 - p_draw_base). "
            "Параметр p_draw_base брать из исторических данных лиги (например, 0.28 для "
            "Английской Премьер-Лиги, 0.30 для колумбийских лиг). Это скорректирует Kelly "
            "вниз для матчей с высокой вероятностью ничьей."
        ),
        "priority": "ВЫСОКИЙ",
        "fill": ORANGE_FILL,
    },
    {
        "n":      4,
        "title":  "Ставка против рынка без обоснования (UNDER на O/U 1.5)",
        "detail": (
            "Рынок Puebla/Tigres давал OVER 75.5% (YES=0.755). Мы поставили NO (UNDER) — "
            "то есть против 75%-й вероятности по рынку. Статистика: ~72-75% матчей "
            "Лиги MX завершаются с тоталом > 1.5 голов. Без sharp данных ставить "
            "против рынка с таким весом = сознательное принятие убытка."
        ),
        "fix": (
            "Добавить правило в signals.py: ЗАПРЕЩЕНО ставить NO, если YES > 0.65 "
            "(рынок сильно перевешен). Для O/U ставок — сравнивать с исторической "
            "статистикой тоталов для конкретной лиги. Добавить модуль data/league_stats.py "
            "с базой средних тоталов по лигам."
        ),
        "priority": "ВЫСОКИЙ",
        "fill": ORANGE_FILL,
    },
    {
        "n":      5,
        "title":  "Kelly Criterion применён без реального edge → завышенные ставки",
        "detail": (
            "Kelly рассчитывался с edge_est=5% (произвольная оценка). "
            "При отсутствии реального edge, Kelly-ставка должна быть 0 — ставить запрещено. "
            "Результат: поставили $30 на Puebla/Tigres (максимально допустимое), хотя "
            "реального преимущества не было. Kelly без edge = гарантированный убыток."
        ),
        "fix": (
            "В kelly.py: edge_est заменить на реально рассчитанный edge из сравнения "
            "с Pinnacle. Если edge <= 0 — bet_size = 0 автоматически. "
            "Добавить minimum_edge = 0.03 (3%) в config.py как обязательный фильтр. "
            "Ставки размещать ТОЛЬКО если Pinnacle_prob > Polymarket_price + 0.03."
        ),
        "priority": "КРИТИЧНО",
        "fill": RED_FILL,
    },
    {
        "n":      6,
        "title":  "Малая выборка — 4 ставки не дают статистической значимости",
        "detail": (
            "4 ставки — слишком мало для выводов. При WR=25% 95%-й доверительный интервал: "
            "[3%, 65%]. Это значит истинный WR может быть и 3% и 65%. "
            "Для статистической значимости нужно минимум 30-50 ставок с положительным EV. "
            "Текущий убыток может быть просто дисперсией."
        ),
        "fix": (
            "Реализовать forward-testing модуль: записывать ВСЕ сигналы с EV>0 в БД "
            "(уже есть snapshots в db.py), не только размещённые ставки. "
            "После 50+ сигналов — анализировать реальный win rate vs. ожидаемый. "
            "Добавить в backtest.py расчёт Sharpe Ratio и графиков доходности."
        ),
        "priority": "СРЕДНИЙ",
        "fill": YELLOW_FILL,
    },
    {
        "n":      7,
        "title":  "Нет фильтра по лиге / типу рынка",
        "detail": (
            "Мы ставили на Liga BetPlay (Колумбия), Liga MX (Мексика), Ligue 1 (Франция). "
            "Эти рынки на Polymarket имеют разное качество ценообразования. "
            "Малоизвестные лиги менее эффективны, но и данных Pinnacle по ним меньше. "
            "Смешивать Tier-1 (EPL, La Liga) и Tier-3 лиги нельзя — разный уровень информации."
        ),
        "fix": (
            "В config.py добавить ALLOWED_LEAGUES список: только EPL, La Liga, Bundesliga, "
            "Serie A, Ligue 1, Champions League, MLS. Liga BetPlay, Liga MX — в TIER2_LEAGUES "
            "с уменьшенным Kelly (0.1 вместо 0.25). Добавить league_detector() в signals.py."
        ),
        "priority": "ВЫСОКИЙ",
        "fill": ORANGE_FILL,
    },
    {
        "n":      8,
        "title":  "Нет проверки времени до начала матча",
        "detail": (
            "Ставки размещались без учёта времени до матча. Оптимальное окно: "
            "за 2-6 часов до старта, когда линия стабилизировалась и известен состав. "
            "Ставить за >24ч — риск смены состава, погоды, дисквалификаций. "
            "Ставить за <30мин — риск резкого движения цены (line movement)."
        ),
        "fix": (
            "В executor.py добавить проверку: if hours_to_start < 2 or hours_to_start > 8: skip. "
            "Подключить ротацию составов: Football-data.org API отдаёт lineups за 1ч до игры. "
            "Если ключевые игроки не играют — отменить ставку или уменьшить размер."
        ),
        "priority": "ВЫСОКИЙ",
        "fill": ORANGE_FILL,
    },
    {
        "n":      9,
        "title":  "Отсутствие стоп-лосса и дневного лимита потерь",
        "detail": (
            "За одну серию ставок потеряно $57.73 (-$10, -$17.73, -$30) прежде чем "
            "получили выигрыш $47.18. Если бы Kelly продолжил генерировать сигналы — "
            "мог бы получиться drawdown в 10-15% bankroll без ограничителей. "
            "RiskManager.check() существует, но daily_loss_limit не активен."
        ),
        "fix": (
            "Активировать в config.py: DAILY_LOSS_LIMIT=50 (5% от bankroll $1000). "
            "После потери $50 в день — полная остановка до следующего дня. "
            "Добавить MAX_CONSECUTIVE_LOSSES=3: после 3 проигрышей подряд — пауза 24ч "
            "для пересмотра стратегии. Всё это уже есть в risk.py — нужно включить."
        ),
        "priority": "ВЫСОКИЙ",
        "fill": ORANGE_FILL,
    },
    {
        "n":      10,
        "title":  "Нет backtesting на исторических данных перед реальными ставками",
        "detail": (
            "Стратегия запущена сразу в 'production' без валидации на истории. "
            "Стандартная практика: сначала backtest на 6-12 месяцев исторических данных, "
            "убедиться в положительном EV, потом forward-test (бумажные ставки 1-3 месяца), "
            "и только затем — реальные деньги. Мы пропустили оба этапа."
        ),
        "fix": (
            "Добавить модуль backtest/historical.py: скачать исторические odds с "
            "the-odds-api.com (платно) или odds-api.io. Симулировать стратегию на "
            "прошлых данных. Добавить в README раздел 'Требования к backtesting'. "
            "Текущий forward-test (db.py snapshots) — продолжать минимум 2 месяца "
            "перед увеличением ставок."
        ),
        "priority": "СРЕДНИЙ",
        "fill": YELLOW_FILL,
    },
]

print()
for r in REASONS:
    print(f"\n[{r['n']:02d}] {r['priority']} — {r['title']}")
    print(f"     Причина: {r['detail'][:120]}...")
    print(f"     Решение: {r['fix'][:120]}...")

# ════════════════════════════════════════════════════════════
# ЗАПИСЫВАЕМ В EXCEL — новая вкладка АНАЛИЗ ОШИБОК
# ════════════════════════════════════════════════════════════

wb = load_workbook(EXCEL_PATH)

# Удаляем старую вкладку если есть
if "Анализ ошибок" in wb.sheetnames:
    del wb["Анализ ошибок"]

ws = wb.create_sheet("Анализ ошибок")
wb.active = ws

# ── Заголовок ─────────────────────────────────────────────
ws.merge_cells("A1:G1")
c = ws.cell(row=1, column=1,
    value=f"АНАЛИЗ ОТРИЦАТЕЛЬНОГО РЕЗУЛЬТАТА — {NOW.strftime('%d.%m.%Y')}")
c.font      = Font(bold=True, color="FFFFFF", size=14)
c.fill      = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
c.alignment = CENTER
ws.row_dimensions[1].height = 35

# ── Сводка результатов ────────────────────────────────────
ws.merge_cells("A2:G2")
ws.cell(row=2, column=1,
    value=f"Итого: {len(bets)} ставок  |  Побед: {len(won_bets)}  Проигрышей: {len(lost_bets)}  "
          f"|  WR: {len(won_bets)/len(bets)*100:.1f}%  |  P&L: ${total_pnl:.2f}  |  ROI: {roi:.1f}%"
).font = Font(bold=True, size=11)
ws.cell(row=2, column=1).fill = RED_FILL
ws.cell(row=2, column=1).alignment = CENTER
ws.row_dimensions[2].height = 22

# ── Таблица ставок ────────────────────────────────────────
ws.row_dimensions[3].height = 5
bet_headers = ["Ставка","Направление","Цена","Сумма $","P&L $","Итог рынка","Ошибки"]
for ci, h in enumerate(bet_headers, 1):
    c = ws.cell(row=4, column=ci, value=h)
    c.font = Font(bold=True, color="FFFFFF", size=10)
    c.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    c.alignment = CENTER
    c.border = THIN
ws.row_dimensions[4].height = 24

for bi, b in enumerate(bets, 5):
    fill = GREEN_FILL if b["won"] else RED_FILL
    ws.cell(row=bi, column=1, value=b["q"]).alignment = WRAP
    ws.cell(row=bi, column=2, value=b["side"]).alignment = CENTER
    ws.cell(row=bi, column=3, value=b["price"]).number_format = "0.000"
    ws.cell(row=bi, column=3).alignment = CENTER
    ws.cell(row=bi, column=4, value=b["bet"]).number_format = '#,##0.00'
    c = ws.cell(row=bi, column=5, value=b["pnl"])
    c.number_format = '+#,##0.00;-#,##0.00'
    c.fill = fill
    ws.cell(row=bi, column=6, value=b["outcome"]).fill = fill
    ws.cell(row=bi, column=6).alignment = CENTER
    errs = "; ".join(b["errors"]) if b["errors"] else "Нет ошибок"
    ws.cell(row=bi, column=7, value=errs).alignment = WRAP
    ws.cell(row=bi, column=7).fill = RED_FILL if b["errors"] else GREEN_FILL
    for ci in range(1, 8):
        ws.cell(row=bi, column=ci).border = THIN
    ws.row_dimensions[bi].height = 36
    if not b["errors"]:
        for ci in range(1, 7):
            ws.cell(row=bi, column=ci).fill = GREEN_FILL

ws.row_dimensions[9].height = 10  # разделитель

# ── 10 причин ────────────────────────────────────────────
row = 10
# Заголовки таблицы причин
reason_headers = ["№","Приоритет","Причина отрицательного результата","Детальный анализ","Что добавить в проект","Статус"]
col_widths_r   = [4,  12,         38,                                  52,                 52,                      12]
for ci, h in enumerate(reason_headers, 1):
    c = ws.cell(row=row, column=ci, value=h)
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    c.alignment = CENTER
    c.border = THIN
ws.row_dimensions[row].height = 28
row += 1

for r in REASONS:
    c_num  = ws.cell(row=row, column=1, value=r["n"])
    c_num.alignment = CENTER
    c_num.fill  = r["fill"]
    c_num.font  = Font(bold=True, size=12)

    c_pri = ws.cell(row=row, column=2, value=r["priority"])
    c_pri.alignment = CENTER
    c_pri.fill  = r["fill"]
    c_pri.font  = Font(bold=True, size=10)

    ws.cell(row=row, column=3, value=r["title"]).font = Font(bold=True, size=10)
    ws.cell(row=row, column=3).fill = r["fill"]
    ws.cell(row=row, column=3).alignment = WRAP

    ws.cell(row=row, column=4, value=r["detail"]).alignment = WRAP

    ws.cell(row=row, column=5, value=r["fix"]).alignment = WRAP
    ws.cell(row=row, column=5).fill = GREEN_FILL

    status_map = {"КРИТИЧНО": "Срочно", "ВЫСОКИЙ": "Планируем", "СРЕДНИЙ": "В очереди"}
    c_st = ws.cell(row=row, column=6, value=status_map.get(r["priority"],"?"))
    c_st.alignment = CENTER
    c_st.fill = r["fill"]
    c_st.font = Font(bold=True, size=10)

    for ci in range(1, 7):
        ws.cell(row=row, column=ci).border = THIN
    ws.row_dimensions[row].height = 80
    row += 1

# ── Итоговый план действий ───────────────────────────────
row += 1
ws.merge_cells(f"A{row}:F{row}")
ws.cell(row=row, column=1,
    value="ПЛАН НЕМЕДЛЕННЫХ УЛУЧШЕНИЙ (ПРИОРИТЕТ 1-3)").font = Font(bold=True, color="FFFFFF", size=12)
ws.cell(row=row, column=1).fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
ws.cell(row=row, column=1).alignment = CENTER
ws.row_dimensions[row].height = 28
row += 1

action_items = [
    ("ШАГ 1 — НЕМЕДЛЕННО",
     "Подключить The Odds API в odds_api.py (ключ уже есть). "
     "Включить фильтр: ставить ТОЛЬКО если Pinnacle_prob > Polymarket_price + 0.03.",
     "КРИТИЧНО", RED_FILL),
    ("ШАГ 2 — НЕМЕДЛЕННО",
     "Включить DAILY_LOSS_LIMIT=50 и MAX_CONSECUTIVE_LOSSES=3 в config.py. "
     "Risk manager уже написан — нужно активировать.",
     "КРИТИЧНО", RED_FILL),
    ("ШАГ 3 — НЕМЕДЛЕННО",
     "Отключить автоставки на UNDER если рынок YES > 0.65. "
     "Добавить в signals.py: if yes_price > 0.65 and side=='NO': skip.",
     "КРИТИЧНО", RED_FILL),
    ("ШАГ 4 — ЭТА НЕДЕЛЯ",
     "Добавить league whitelist: только EPL, La Liga, Bundesliga, Serie A, Ligue 1. "
     "Liga MX / Liga BetPlay — половинный Kelly (0.10).",
     "ВЫСОКИЙ", ORANGE_FILL),
    ("ШАГ 5 — ЭТА НЕДЕЛЯ",
     "Добавить draw_probability_correction в kelly.py. "
     "Для матчей: p_adjusted = p_yes / (1 - 0.27). Тогда Kelly будет точнее.",
     "ВЫСОКИЙ", ORANGE_FILL),
    ("ШАГ 6 — ЭТОТ МЕСЯЦ",
     "Реализовать forward-test на 50+ сигналах без реальных ставок. "
     "Только после подтверждения WR > 55% — включать реальные деньги.",
     "СРЕДНИЙ", YELLOW_FILL),
]

for step, desc, prio, fill in action_items:
    ws.cell(row=row, column=1, value=step).font = Font(bold=True, size=10)
    ws.cell(row=row, column=1).fill = fill
    ws.cell(row=row, column=1).alignment = CENTER
    ws.merge_cells(f"B{row}:E{row}")
    ws.cell(row=row, column=2, value=desc).alignment = WRAP
    ws.cell(row=row, column=2).border = THIN
    ws.cell(row=row, column=6, value=prio).alignment = CENTER
    ws.cell(row=row, column=6).fill = fill
    ws.cell(row=row, column=6).font = Font(bold=True, size=10)
    ws.cell(row=row, column=6).border = THIN
    ws.cell(row=row, column=1).border = THIN
    ws.row_dimensions[row].height = 55
    row += 1

# ── Ширина столбцов ───────────────────────────────────────
for col, width in zip("ABCDEF", [4, 12, 38, 52, 52, 12]):
    ws.column_dimensions[get_column_letter(["A","B","C","D","E","F"].index(col)+1)].width = width

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 35
ws.column_dimensions["D"].width = 48
ws.column_dimensions["E"].width = 48
ws.column_dimensions["F"].width = 13

wb.save(EXCEL_PATH)
print(f"\n  Excel обновлён: вкладка 'Анализ ошибок' создана")
print(f"  Файл: {EXCEL_PATH}")
