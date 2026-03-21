"""Обновить прогнозы из закэшированного файла (если API недоступен)."""
import json, sys, io
from datetime import datetime, timezone, timedelta
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CACHE_DIR = Path(r"C:\Users\Lomov\.cursor\projects\c-Users-Lomov-Documents-Arduino-libraries-M5Unified-src-utility\agent-tools")
EXCEL_PATH = Path(__file__).parent / "polymarket_рекомендации.xlsx"
NOW = datetime.now(timezone.utc)
THIN = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
ZEBRA = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

SPORT_KW = ["win","beat"," vs ","o/u","over","under","nba","nhl","ufc","fc ","afc","arsenal","chelsea","city","dodgers","england","leverkusen","bayer"]
def cat(q): return "Спорт" if any(w in q.lower() for w in SPORT_KW) else "Другое"
def rec(cat, liq):
    if cat=="Спорт" and liq>=10_000: return "АНАЛИЗИРОВАТЬ", GREEN
    if cat=="Спорт" and liq>=2_000: return "ОСТОРОЖНО", YELLOW
    if cat=="Спорт": return "НЕТ ЛИК-ТИ", ORANGE
    return "ПРОПУСТИТЬ", RED
WHY = {"АНАЛИЗИРОВАТЬ":"Спорт + высокая ликв.","ОСТОРОЖНО":"Спорт, умеренная ликв.","НЕТ ЛИК-ТИ":"Низкая ликв.","ПРОПУСТИТЬ":"Нет бенчмарка."}

data = []
for fp in CACHE_DIR.glob("*.txt"):
    try:
        raw = json.load(open(fp, encoding="utf-8"))
        data.extend(raw if isinstance(raw, list) else [])
    except Exception:
        pass
seen = set()
dedup = []
for m in data:
    cid = m.get("conditionId") or m.get("id", "")
    if cid and cid not in seen:
        seen.add(cid)
        dedup.append(m)
data = dedup
parsed = []
for m in data:
    p = m.get("outcomePrices","[]")
    if isinstance(p, str): p = json.loads(p) if p else []
    if len(p) < 2: continue
    liq = float(m.get("liquidity",0) or 0)
    if liq < 500: continue
    yes = float(p[0])
    if not (0.04 < yes < 0.96): continue
    end = m.get("endDate","")
    end_dt = datetime.fromisoformat(end.replace("Z","+00:00")) if end else None
    if end_dt and end_dt < NOW: continue
    c = cat(m.get("question",""))
    r, fill = rec(c, liq)
    pri = 0 if r=="АНАЛИЗИРОВАТЬ" else 1 if r=="ОСТОРОЖНО" else 9
    parsed.append({"q":m.get("question",""),"cat":c,"yes":yes,"no":float(p[1]),"liq":liq,"vol24":float(m.get("volume24hr",0) or 0),"end_dt":end_dt,"rec":r,"rfill":fill,"pri":pri})
parsed.sort(key=lambda x: (x["pri"], -x["liq"]))
display = parsed[:30]

wb = load_workbook(EXCEL_PATH)
ws = wb["Рекомендации"]
for r in range(2, 62):
    for c in range(1, 12): ws.cell(row=r, column=c).value = None
h = ["№","Рынок","Категория","ДА","НЕТ","Ликв $","Vol24h","Конец","Рекомендация","Обоснование","Проверить"]
for c, x in enumerate(h, 1): ws.cell(row=1, column=c, value=x)
for i, m in enumerate(display, 1):
    r = i + 1
    ws.cell(row=r, column=1, value=i)
    ws.cell(row=r, column=2, value=m["q"])
    ws.cell(row=r, column=3, value=m["cat"])
    ws.cell(row=r, column=4, value=m["yes"])
    ws.cell(row=r, column=5, value=m["no"])
    ws.cell(row=r, column=6, value=m["liq"])
    ws.cell(row=r, column=7, value=m["vol24"])
    ws.cell(row=r, column=8, value=m["end_dt"].strftime("%Y-%m-%d %H:%M") if m["end_dt"] else "")
    ws.cell(row=r, column=9, value=m["rec"]).fill = m["rfill"]
    ws.cell(row=r, column=10, value=WHY.get(m["rec"],""))
    ws.cell(row=r, column=11, value=(m["end_dt"]+timedelta(hours=2)).strftime("%Y-%m-%d %H:%M") if m["end_dt"] else "")
    for c in range(1, 12): ws.cell(row=r, column=c).fill = ZEBRA if i%2==0 else PatternFill()
    ws.cell(row=r, column=9).fill = m["rfill"]

ws_stats = wb["Статистика"]
ws_stats.cell(row=7, column=1, value="Обновлено")
ws_stats.cell(row=7, column=2, value=NOW.strftime("%Y-%m-%d %H:%M UTC"))
wb.save(EXCEL_PATH)
print(f"Обновлено {len(display)} рекомендаций из кэша. Excel: {EXCEL_PATH}")
for m in [x for x in display if x["rec"] in ("АНАЛИЗИРОВАТЬ","ОСТОРОЖНО")][:10]:
    print(f"  [{m['rec']}] {m['q'][:55]} ДА={m['yes']:.3f} ${m['liq']:,.0f}")
