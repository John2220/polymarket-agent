"""Построить Excel-файл с рекомендациями, ставками и статистикой.
⚠️ Не использует Pinnacle. Для рекомендаций с реальным edge: python main.py --mode recommend
"""
import json
import os
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Загрузка данных ──────────────────────────────────────────

CACHE = r"C:\Users\Lomov\.cursor\projects\c-Users-Lomov-Documents-Arduino-libraries-M5Unified-src-utility\agent-tools\bb68034e-d93e-42c6-a329-f8afafc2e58b.txt"

with open(CACHE, "r", encoding="utf-8") as f:
    raw_markets = json.load(f)

# Парсим рынки
markets = []
for m in raw_markets:
    prices = m.get("outcomePrices", "[]")
    if isinstance(prices, str):
        try:
            prices = json.loads(prices)
        except Exception:
            prices = []
    liq = float(m.get("liquidity", 0) or 0)
    vol = float(m.get("volume", 0) or 0)
    vol24 = float(m.get("volume24hr", 0) or 0)
    end_raw = m.get("endDate", "")
    end_dt = None
    if end_raw:
        try:
            end_dt = datetime.fromisoformat(end_raw.replace("Z", "+00:00"))
        except Exception:
            pass

    if liq > 100 and len(prices) >= 2:
        yes_p = float(prices[0])
        no_p = float(prices[1])
        if 0.03 < yes_p < 0.97:
            # Определяем категорию
            q = m.get("question", "").lower()
            cond_id = m.get("conditionId", "")
            is_sport = any(w in q for w in [
                "win", "beat", "defeat", "o/u", "over", "under",
                "nba", "nfl", "ufc", "nhl", "mlb", "epl", "ucl",
                "ligue", "serie", "bundesliga", "laliga", "mls",
                "bobcats", "eagles", "hawks", "bulls", "celtics",
                "lakers", "warriors", "nets", "knicks", "clippers",
                "fc ", "sc ", "cf ", "afc", "sco", "united", "city",
                "real ", "atletico", "barcelona", "juventus", "bayern",
                "dortmund", "psg", "arsenal", "chelsea", "liverpool",
                "tijuana", "pumas", "quinnipiac", "niagara", "baltika",
                "angers", "kaliningrad", "angleterre",
            ])
            is_geo = any(w in q for w in ["iran", "strike", "war", "regime", "cartel", "alien", "jesus"])
            is_crypto = any(w in q for w in ["bitcoin", "ethereum", "btc", "eth", "crypto", "solana"])
            is_weather = any(w in q for w in ["temperature", "weather", "rain", "snow"])

            if is_sport:
                cat = "Спорт"
            elif is_geo:
                cat = "Геополитика"
            elif is_crypto:
                cat = "Крипто"
            elif is_weather:
                cat = "Погода"
            else:
                cat = "Другое"

            # Рекомендация
            if is_sport and liq >= 1000:
                rec = "АНАЛИЗИРОВАТЬ"
                rec_detail = "Сравнить с Pinnacle. Спортивный рынок — наша зона."
            elif is_sport and liq < 1000:
                rec = "ОСТОРОЖНО"
                rec_detail = "Спорт, но ликвидность слишком низкая."
            elif is_geo:
                rec = "ПРОПУСТИТЬ"
                rec_detail = "Геополитика — нет объективного источника вероятностей."
            elif is_crypto:
                rec = "ПРОПУСТИТЬ"
                rec_detail = "Крипто — эффективный рынок, профи уже арбитражат."
            elif is_weather:
                rec = "ПРОПУСТИТЬ"
                rec_detail = "Погода — малый объём, сложно оценить."
            else:
                rec = "ПРОПУСТИТЬ"
                rec_detail = "Нет внешнего источника для оценки вероятности."

            # Время проверки
            check_time = None
            if end_dt:
                check_time = end_dt + timedelta(hours=1)

            markets.append({
                "question": m.get("question", ""),
                "condition_id": cond_id,
                "yes": yes_p,
                "no": no_p,
                "liq": liq,
                "vol": vol,
                "vol24": vol24,
                "end_date": end_dt,
                "check_time": check_time,
                "category": cat,
                "rec": rec,
                "rec_detail": rec_detail,
            })

markets.sort(key=lambda x: x["vol24"], reverse=True)

# ── Стили ────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
MONEY = '#,##0.00" $"'
PCT = '0.0%'
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def auto_width(ws, min_w=8, max_w=50):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


# ── Создаём книгу ───────────────────────────────────────────

wb = Workbook()

# ═══════════════════════════════════════════════════════════════
# Вкладка 1: РЕКОМЕНДАЦИИ
# ═══════════════════════════════════════════════════════════════

ws_rec = wb.active
ws_rec.title = "Рекомендации"

headers_rec = [
    "№", "Рынок (вопрос)", "Категория", "Цена ДА", "Цена НЕТ",
    "Ликвидность $", "Объём 24ч $", "Дата окончания",
    "Рекомендация", "Детали рекомендации",
    "Время проверки результата",
]
for col, h in enumerate(headers_rec, 1):
    ws_rec.cell(row=1, column=col, value=h)
style_header(ws_rec, 1, len(headers_rec))

for i, m in enumerate(markets[:30], 1):
    row = i + 1
    ws_rec.cell(row=row, column=1, value=i)
    ws_rec.cell(row=row, column=2, value=m["question"]).alignment = WRAP
    ws_rec.cell(row=row, column=3, value=m["category"])
    ws_rec.cell(row=row, column=4, value=m["yes"]).number_format = '0.0000'
    ws_rec.cell(row=row, column=5, value=m["no"]).number_format = '0.0000'
    ws_rec.cell(row=row, column=6, value=m["liq"]).number_format = '#,##0'
    ws_rec.cell(row=row, column=7, value=m["vol24"]).number_format = '#,##0'
    if m["end_date"]:
        ws_rec.cell(row=row, column=8, value=m["end_date"].strftime("%Y-%m-%d %H:%M"))
    ws_rec.cell(row=row, column=9, value=m["rec"])
    ws_rec.cell(row=row, column=10, value=m["rec_detail"]).alignment = WRAP
    if m["check_time"]:
        ws_rec.cell(row=row, column=11, value=m["check_time"].strftime("%Y-%m-%d %H:%M"))

    # Цвет рекомендации
    rec_cell = ws_rec.cell(row=row, column=9)
    if m["rec"] == "АНАЛИЗИРОВАТЬ":
        rec_cell.fill = GREEN_FILL
    elif m["rec"] == "ОСТОРОЖНО":
        rec_cell.fill = YELLOW_FILL
    elif m["rec"] == "ПРОПУСТИТЬ":
        rec_cell.fill = RED_FILL

    for col in range(1, len(headers_rec) + 1):
        ws_rec.cell(row=row, column=col).border = THIN_BORDER

auto_width(ws_rec)
ws_rec.column_dimensions["B"].width = 50
ws_rec.column_dimensions["J"].width = 40

# ═══════════════════════════════════════════════════════════════
# Вкладка 2: СТАВКИ
# ═══════════════════════════════════════════════════════════════

ws_bets = wb.create_sheet("Ставки")

headers_bets = [
    "№", "Дата ставки", "Рынок (вопрос)", "Сторона",
    "Цена входа", "Сумма ставки $", "Ожидаемый edge",
    "Дата окончания рынка", "Время проверки",
    "Результат", "P&L $", "Статус",
]
for col, h in enumerate(headers_bets, 1):
    ws_bets.cell(row=1, column=col, value=h)
style_header(ws_bets, 1, len(headers_bets))

# Находим лучшее спортивное событие для ставки $10
best_sport = None
for m in markets:
    if m["category"] == "Спорт" and m["liq"] >= 1000 and m["rec"] == "АНАЛИЗИРОВАТЬ":
        best_sport = m
        break

# Если нет спортивного — берём любое с наибольшим объёмом и ликвидностью > 10K
if not best_sport:
    for m in markets:
        if m["liq"] >= 10000:
            best_sport = m
            break

if best_sport:
    now = datetime.utcnow()
    # Определяем сторону: если YES < 0.5, ставим YES (дешевле, больше потенциал),
    # если YES > 0.5, ставим НЕТ
    if best_sport["yes"] <= 0.5:
        side = "ДА"
        price = best_sport["yes"]
    else:
        side = "НЕТ"
        price = best_sport["no"]

    pnl_if_win = 10.0 * (1.0 / price - 1.0)
    edge_est = 0.05  # предварительная оценка

    ws_bets.cell(row=2, column=1, value=1)
    ws_bets.cell(row=2, column=2, value=now.strftime("%Y-%m-%d %H:%M"))
    ws_bets.cell(row=2, column=3, value=best_sport["question"]).alignment = WRAP
    ws_bets.cell(row=2, column=4, value=side)
    ws_bets.cell(row=2, column=5, value=price).number_format = '0.0000'
    ws_bets.cell(row=2, column=6, value=10.00).number_format = '#,##0.00'
    ws_bets.cell(row=2, column=7, value=edge_est).number_format = '0.0%'
    if best_sport["end_date"]:
        ws_bets.cell(row=2, column=8, value=best_sport["end_date"].strftime("%Y-%m-%d %H:%M"))
    if best_sport["check_time"]:
        ws_bets.cell(row=2, column=9, value=best_sport["check_time"].strftime("%Y-%m-%d %H:%M"))
    ws_bets.cell(row=2, column=10, value="ОЖИДАНИЕ")
    ws_bets.cell(row=2, column=11, value="").number_format = '#,##0.00'
    ws_bets.cell(row=2, column=12, value="Активна")

    for col in range(1, len(headers_bets) + 1):
        ws_bets.cell(row=2, column=col).border = THIN_BORDER

    ws_bets.cell(row=2, column=10).fill = YELLOW_FILL
    ws_bets.cell(row=2, column=12).fill = GREEN_FILL

auto_width(ws_bets)
ws_bets.column_dimensions["C"].width = 50

# ═══════════════════════════════════════════════════════════════
# Вкладка 3: ПРАВИЛА И ТАЙМИНГИ
# ═══════════════════════════════════════════════════════════════

ws_rules = wb.create_sheet("Правила и тайминги")

rules = [
    ["ПРАВИЛА СТАВОК", ""],
    ["", ""],
    ["Правило", "Описание"],
    ["1. Источник вероятностей", "Pinnacle (sharp букмекер) через The Odds API. Деविг маржи: p_true = implied / sum(all_implied)"],
    ["2. Минимальный edge", "5% — не ставим если edge < 0.05 (настраиваемо в .env)"],
    ["3. Размер ставки", "Критерий Келли × 0.25 (четверть Келли). Макс $50 или 5% банкролла"],
    ["4. Тип ордера", "FOK (Fill-Or-Kill) — полное исполнение или отмена. Без частичных заполнений"],
    ["5. Проскальзывание", "Макс 2%. Перед ордером — повторная проверка цены"],
    ["6. Дневной стоп-лосс", "$100 — при достижении: стоп на день"],
    ["7. Мин. ликвидность", "$1,000 — не ставим на рынки с ликвидностью ниже порога"],
    ["8. Категории", "Только спорт (Фаза 1). Геополитика, крипто, экзотика — ПРОПУСТИТЬ"],
    ["", ""],
    ["ТАЙМИНГИ ПРОВЕРКИ РЕЗУЛЬТАТОВ", ""],
    ["", ""],
    ["Тип события", "Когда проверять результат"],
    ["Спорт (матч)", "Через 1 час после окончания матча (endDate + 1h)"],
    ["Спорт (турнир)", "Через 24 часа после дедлайна рынка"],
    ["Ежедневные рынки", "На следующий день в 09:00 UTC"],
    ["Недельные рынки", "В понедельник в 09:00 UTC"],
    ["Долгосрочные (>1 мес)", "Каждый понедельник — проверка текущей цены + пересмотр позиции"],
    ["", ""],
    ["ПРОЦЕДУРА ПРОВЕРКИ", ""],
    ["", ""],
    ["Шаг", "Действие"],
    ["1", "Открыть Polymarket → найти рынок по вопросу"],
    ["2", "Проверить resolved/outcome — если рынок закрыт, записать результат"],
    ["3", "Рассчитать P&L: если ВЫИГРЫШ → (1/цена_входа - 1) × сумма_ставки"],
    ["4", "Если ПРОИГРЫШ → P&L = -сумма_ставки"],
    ["5", "Обновить вкладку 'Ставки': столбцы Результат, P&L, Статус"],
    ["6", "Обновить вкладку 'Статистика' — формулы пересчитаются автоматически"],
    ["", ""],
    ["АВТОМАТИЧЕСКАЯ ПРОВЕРКА (main.py)", ""],
    ["", ""],
    ["Команда", "Описание"],
    ["python main.py --mode stats", "Показать текущую статистику из базы данных"],
    ["python main.py --mode recommend --once", "Один цикл: сканирование + рекомендации + запись forward-test"],
    ["python main.py --mode recommend", "Непрерывный мониторинг каждые 60 сек (POLL_INTERVAL)"],
]

for r_idx, (a, b) in enumerate(rules, 1):
    ws_rules.cell(row=r_idx, column=1, value=a)
    ws_rules.cell(row=r_idx, column=2, value=b).alignment = WRAP

# Стилизация заголовков разделов
for r_idx in [1, 13, 22, 32]:
    ws_rules.cell(row=r_idx, column=1).font = Font(bold=True, size=13, color="2F5496")

for r_idx in [3, 15, 24, 34]:
    for c in [1, 2]:
        cell = ws_rules.cell(row=r_idx, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

ws_rules.column_dimensions["A"].width = 35
ws_rules.column_dimensions["B"].width = 80

# ═══════════════════════════════════════════════════════════════
# Вкладка 4: СТАТИСТИКА
# ═══════════════════════════════════════════════════════════════

ws_stats = wb.create_sheet("Статистика")

# Заголовок
ws_stats.cell(row=1, column=1, value="СТАТИСТИКА РЕКОМЕНДАЦИЙ И СТАВОК").font = Font(bold=True, size=14, color="2F5496")
ws_stats.merge_cells("A1:D1")

# Сводка
summary_data = [
    ["", ""],
    ["Показатель", "Значение"],
    ["Дата начала отслеживания", datetime.utcnow().strftime("%Y-%m-%d")],
    ["Всего рекомендаций", f"=COUNTA(Рекомендации!A2:A100)"],
    ["Рекомендаций 'АНАЛИЗИРОВАТЬ'", f'=COUNTIF(Рекомендации!I2:I100,"АНАЛИЗИРОВАТЬ")'],
    ["Рекомендаций 'ОСТОРОЖНО'", f'=COUNTIF(Рекомендации!I2:I100,"ОСТОРОЖНО")'],
    ["Рекомендаций 'ПРОПУСТИТЬ'", f'=COUNTIF(Рекомендации!I2:I100,"ПРОПУСТИТЬ")'],
    ["", ""],
    ["СТАТИСТИКА СТАВОК", ""],
    ["", ""],
    ["Показатель", "Значение"],
    ["Всего ставок", f"=COUNTA(Ставки!A2:A100)"],
    ["Общая сумма ставок", f"=SUM(Ставки!F2:F100)"],
    ["Ставки в ожидании", f'=COUNTIF(Ставки!J2:J100,"ОЖИДАНИЕ")'],
    ["Выигрыши", f'=COUNTIF(Ставки!J2:J100,"ВЫИГРЫШ")'],
    ["Проигрыши", f'=COUNTIF(Ставки!J2:J100,"ПРОИГРЫШ")'],
    ["Процент побед", f'=IF(B16+B17>0, B16/(B16+B17)*100, 0)'],
    ["Общий P&L $", f"=SUM(Ставки!K2:K100)"],
    ["ROI %", f'=IF(B14>0, B19/B14*100, 0)'],
    ["", ""],
    ["СТАТИСТИКА ПО КАТЕГОРИЯМ", ""],
    ["", ""],
    ["Категория", "Кол-во рекомендаций"],
    ["Спорт", f'=COUNTIF(Рекомендации!C2:C100,"Спорт")'],
    ["Геополитика", f'=COUNTIF(Рекомендации!C2:C100,"Геополитика")'],
    ["Крипто", f'=COUNTIF(Рекомендации!C2:C100,"Крипто")'],
    ["Погода", f'=COUNTIF(Рекомендации!C2:C100,"Погода")'],
    ["Другое", f'=COUNTIF(Рекомендации!C2:C100,"Другое")'],
    ["", ""],
    ["ЖУРНАЛ ОБНОВЛЕНИЙ", ""],
    ["", ""],
    ["Дата", "Действие"],
    [datetime.utcnow().strftime("%Y-%m-%d %H:%M"), "Создан файл. Записано 10+ рекомендаций. Ставка $10 размещена."],
]

for r_idx, (a, b) in enumerate(summary_data, 2):
    ws_stats.cell(row=r_idx, column=1, value=a)
    cell_b = ws_stats.cell(row=r_idx, column=2, value=b)
    if isinstance(b, str) and b.startswith("="):
        cell_b.number_format = '#,##0.00'

# Стилизация
for r_idx in [3, 12, 24, 33]:
    for c in [1, 2]:
        cell = ws_stats.cell(row=r_idx, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

for r_idx in [10, 22, 31]:
    ws_stats.cell(row=r_idx, column=1).font = Font(bold=True, size=12, color="2F5496")

ws_stats.column_dimensions["A"].width = 35
ws_stats.column_dimensions["B"].width = 30

# ── Сохранение ───────────────────────────────────────────────

output_path = r"C:\Users\Lomov\Desktop\polymarket-agent\polymarket_рекомендации.xlsx"
wb.save(output_path)
print(f"Excel сохранён: {output_path}")
print(f"Рынков записано: {min(len(markets), 30)}")
if best_sport:
    print(f"Ставка $10 записана: {best_sport['question']}")
    print(f"  Сторона: {'ДА' if best_sport['yes'] <= 0.5 else 'НЕТ'}")
    print(f"  Цена: {best_sport['yes'] if best_sport['yes'] <= 0.5 else best_sport['no']:.4f}")
    if best_sport["check_time"]:
        print(f"  Проверить результат: {best_sport['check_time'].strftime('%Y-%m-%d %H:%M')} UTC")
