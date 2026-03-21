"""
Проверка результатов открытых ставок + обновление Excel.
"""
from __future__ import annotations
import json, sys, io
from datetime import datetime, timezone, timedelta
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from core.bet_results import calc_pnl
from core.gamma_resolve import (
    bet_won_for_binary_market,
    binary_outcome_yes_no,
    market_fully_resolved,
)

# ── Файлы данных ─────────────────────────────────────────────
CLOSED_FILE  = r"C:\Users\Lomov\.cursor\projects\c-Users-Lomov-Documents-Arduino-libraries-M5Unified-src-utility\agent-tools\829ae56c-7e23-4760-8198-960a0e7308dc.txt"
ACTIVE_FILES = [
    r"C:\Users\Lomov\.cursor\projects\c-Users-Lomov-Documents-Arduino-libraries-M5Unified-src-utility\agent-tools\f3d1dc97-4e85-49a6-adde-77baf705146d.txt",
    r"C:\Users\Lomov\.cursor\projects\c-Users-Lomov-Documents-Arduino-libraries-M5Unified-src-utility\agent-tools\dc4a98c8-e43e-446e-99be-41e6776e3c4b.txt",
    r"C:\Users\Lomov\.cursor\projects\c-Users-Lomov-Documents-Arduino-libraries-M5Unified-src-utility\agent-tools\c735c5a6-ba47-4065-b627-c6b362a0e199.txt",
]
EXCEL_PATH = Path(r"C:\Users\Lomov\Desktop\polymarket-agent\polymarket_рекомендации.xlsx")
NOW = datetime.now(timezone.utc)

# ── Стили ────────────────────────────────────────────────────
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL  = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
GREEN_FILL   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL     = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL  = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
BLUE_FILL    = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
GRAY_FILL    = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ZEBRA_FILL   = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
CENTER       = Alignment(horizontal="center", vertical="center")
WRAP         = Alignment(wrap_text=True, vertical="top")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

def hdr(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = CENTER; cell.border = THIN_BORDER

# ════════════════════════════════════════════════════════════
# 1. ЗАГРУЖАЕМ ЗАКРЫТЫЕ И АКТИВНЫЕ РЫНКИ
# ════════════════════════════════════════════════════════════

def load_json(path):
    try:
        return json.load(open(path, encoding="utf-8"))
    except Exception as e:
        print(f"  [!] Ошибка: {e}")
        return []

closed_raw  = load_json(CLOSED_FILE)
print(f"  Закрытых рынков загружено: {len(closed_raw)}")

# Индекс закрытых: по conditionId, slug, первым словам вопроса
closed_idx: dict[str, dict] = {}
for m in closed_raw:
    cid  = m.get("conditionId", "")
    slug = m.get("slug", "")
    q    = (m.get("question") or "").lower().strip()
    if cid:  closed_idx[cid]  = m
    if slug: closed_idx[slug] = m
    if q:
        # Индекс по первым 5 словам
        key5 = " ".join(q.split()[:5])
        closed_idx[key5] = m
        # И по первым 4
        key4 = " ".join(q.split()[:4])
        closed_idx[key4] = m

# Загружаем активные
active_raw: list[dict] = []
seen = set()
for fp in ACTIVE_FILES:
    for m in load_json(fp):
        cid = m.get("conditionId","") or m.get("id","")
        if cid in seen: continue
        seen.add(cid)
        active_raw.append(m)
print(f"  Активных рынков загружено: {len(active_raw)}")

# Индекс активных (для обновления цен)
active_idx: dict[str, dict] = {}
for m in active_raw:
    cid  = m.get("conditionId", "")
    slug = m.get("slug", "")
    if cid:  active_idx[cid]  = m
    if slug: active_idx[slug] = m


# ════════════════════════════════════════════════════════════
# 2. ПОИСК РЕЗУЛЬТАТА ПО ВОПРОСУ
# ════════════════════════════════════════════════════════════

def find_closed(question: str, cid: str) -> dict | None:
    """Ищет закрытый рынок с результатом (в т.ч. без флага resolved, по outcomePrices)."""
    # 1. Точный cid
    if cid and cid in closed_idx:
        m = closed_idx[cid]
        if market_fully_resolved(m):
            return m
    # 2. По первым 5 словам
    q = (question or "").lower().strip()
    for n in (5, 4, 3):
        key = " ".join(q.split()[:n])
        if key in closed_idx:
            m = closed_idx[key]
            if market_fully_resolved(m):
                return m
    # 3. Fuzzy: ищем матч по наличию команды в вопросе
    words = [w for w in q.split() if len(w) > 4]
    for key, m in closed_idx.items():
        if not market_fully_resolved(m):
            continue
        mk = str(key).lower()
        if sum(1 for w in words if w in mk) >= 2:
            return m
    return None


def get_current_prices(question: str, cid: str) -> tuple[float | None, float | None]:
    """Возвращает (yes_price, no_price) для активной ставки. Нужно для floating P&L по стороне."""
    m = None
    if cid and cid in active_idx:
        m = active_idx[cid]
    if m is None:
        q = (question or "").lower()
        for n in (5, 4):
            key = " ".join(q.split()[:n])
            if key in active_idx:
                m = active_idx[key]
                break
    if m is None:
        return None, None
    prices = m.get("outcomePrices", "[]")
    if isinstance(prices, str):
        try: prices = json.loads(prices)
        except: return None, None
    if len(prices) >= 2:
        try: return float(prices[0]), float(prices[1])
        except: pass
    if len(prices) >= 1:
        try: y = float(prices[0]); return y, 1.0 - y
        except: pass
    return None, None


# ════════════════════════════════════════════════════════════
# 3. ОБНОВЛЯЕМ ВКЛАДКУ СТАВКИ
# ════════════════════════════════════════════════════════════

wb       = load_workbook(EXCEL_PATH)
ws_bets  = wb["Ставки"]
ws_stats = wb["Статистика"]
ws_rec   = wb["Рекомендации"]

updated   = []
still_open = []
total_pnl  = 0.0
won_cnt = lost_cnt = open_cnt = 0

row = 2
while True:
    q_val   = ws_bets.cell(row=row, column=3).value
    if q_val is None:
        break
    question = str(q_val)
    cid      = str(ws_bets.cell(row=row, column=2).value or "")
    side     = str(ws_bets.cell(row=row, column=4).value or "YES")
    res_cur  = ws_bets.cell(row=row, column=10).value

    try:
        bet_usd = float(str(ws_bets.cell(row=row, column=6).value or 10).replace("$",""))
        price   = float(str(ws_bets.cell(row=row, column=7).value or 0.5).replace("$",""))
    except Exception:
        bet_usd, price = 10.0, 0.5

    # Пропускаем уже закрытые
    if res_cur in ("ВЫИГРЫШ", "ПРОИГРЫШ"):
        try:
            pnl_v = float(str(ws_bets.cell(row=row, column=11).value or 0))
        except:
            pnl_v = 0.0
        total_pnl += pnl_v
        if res_cur == "ВЫИГРЫШ":
            won_cnt += 1
        else:
            lost_cnt += 1
        row += 1
        continue

    # Ищем результат
    closed_m = find_closed(question, cid)

    if closed_m:
        won = bet_won_for_binary_market(side, closed_m)
        if won is None:
            row += 1
            continue

        pnl = calc_pnl(bet_usd, price, won)

        res_ru    = "ВЫИГРЫШ" if won else "ПРОИГРЫШ"
        res_fill  = GREEN_FILL if won else RED_FILL
        pnl_sign  = f"+${pnl:.2f}" if pnl > 0 else f"-${abs(pnl):.2f}"

        ws_bets.cell(row=row, column=10, value=res_ru).fill  = res_fill
        ws_bets.cell(row=row, column=11, value=round(pnl,2)).fill = res_fill
        ws_bets.cell(row=row, column=11).number_format = '#,##0.00'
        ws_bets.cell(row=row, column=12, value="Завершена").fill = GRAY_FILL
        ws_bets.cell(row=row, column=12).alignment = CENTER
        ws_bets.cell(row=row, column=10).alignment = CENTER

        for c in [10,11,12]:
            ws_bets.cell(row=row, column=c).border = THIN_BORDER

        outcome_disp = binary_outcome_yes_no(closed_m) or closed_m.get("outcome", "?")
        updated.append((question[:55], side, res_ru, pnl_sign, outcome_disp))
        total_pnl += pnl
        if won: won_cnt  += 1
        else:   lost_cnt += 1

    else:
        # Ставка ещё активна — обновляем текущую цену (floating P&L)
        # Для NO используем cur_no_price! См. CONTEXT_AND_RULES 2.2
        open_cnt += 1
        yes_p, no_p = get_current_prices(question, cid)
        cur_side_price = (no_p if side.upper() in ("NO", "НЕТ") else yes_p)
        float_pnl = None
        if cur_side_price is not None and price > 0:
            shares = bet_usd / price
            float_val = shares * cur_side_price - bet_usd
            float_pnl = round(float_val, 2)
            ws_bets.cell(row=row, column=11, value=float_pnl)
            ws_bets.cell(row=row, column=11).number_format = '#,##0.00'
            ws_bets.cell(row=row, column=11).fill = \
                GREEN_FILL if float_pnl >= 0 else ORANGE_FILL
            ws_bets.cell(row=row, column=11).border = THIN_BORDER

        end_str = str(ws_bets.cell(row=row, column=14).value or "?")
        still_open.append((question[:55], side, cur_side_price, float_pnl, end_str))

    row += 1

total_closed = won_cnt + lost_cnt
wr = round(won_cnt / total_closed * 100, 1) if total_closed > 0 else 0.0

print(f"\n  Ставок обновлено: {len(updated)}")
print(f"  Ставок ещё активно: {open_cnt}")


# ════════════════════════════════════════════════════════════
# 4. ОБНОВЛЯЕМ ВКЛАДКУ СТАТИСТИКА
# ════════════════════════════════════════════════════════════

stats = {
    2:  ("Всего закрытых ставок", total_closed),
    3:  ("Выигрышей",             won_cnt),
    4:  ("Проигрышей",            lost_cnt),
    5:  ("Win Rate (%)",          wr),
    6:  ("Итоговый P&L ($)",      round(total_pnl, 2)),
    7:  ("Дата обновления",       NOW.strftime("%Y-%m-%d %H:%M UTC")),
    8:  ("Активных ставок",       open_cnt),
}
for rn, (label, val) in stats.items():
    ws_stats.cell(row=rn, column=1, value=label).font = Font(bold=True)
    ws_stats.cell(row=rn, column=2, value=val)
    if rn == 6:
        ws_stats.cell(row=rn, column=2).fill = \
            GREEN_FILL if total_pnl >= 0 else RED_FILL
    for c in [1,2]:
        ws_stats.cell(row=rn, column=c).border = THIN_BORDER

# Журнал
log_row = 12
while ws_stats.cell(row=log_row, column=1).value not in (None, ""):
    log_row += 1
    if log_row > 300: break

log_msg = (
    f"Проверка результатов. "
    f"Закрыто: {len(updated)}, "
    f"Активных: {open_cnt}. "
    f"P&L итого: ${total_pnl:+.2f}, WR={wr}%"
)
ws_stats.cell(row=log_row, column=1, value=NOW.strftime("%Y-%m-%d %H:%M"))
ws_stats.cell(row=log_row, column=2, value=log_msg)
ws_stats.cell(row=log_row, column=2).alignment = WRAP
for c in [1, 2]:
    ws_stats.cell(row=log_row, column=c).border = THIN_BORDER
ws_stats.cell(row=log_row, column=1).fill = \
    GREEN_FILL if total_pnl >= 0 else RED_FILL


# ════════════════════════════════════════════════════════════
# 5. ОБНОВЛЯЕМ ВКЛАДКУ РЕКОМЕНДАЦИИ
# ════════════════════════════════════════════════════════════

SPORT_KW = [
    "win","beat"," vs ","o/u","over","under","total","nba","nhl","nfl",
    "ufc","mlb","epl","ucl","fc "," fc","afc","sc ","tolima","tijuana",
    "puebla","tigres","mazatlan","atletico","pachuca","monterrey",
    "auxerre","strasbourg","lens","metz","gracheva","tagger","tennis",
    "hurricanes","stanley cup","burnley","arsenal","chelsea","liverpool",
    "tottenham","manchester","villa","lakers","celtics","warriors",
    "nuggets","bucks","fight","bout","match",
]

def categorize(q):
    ql = q.lower()
    if any(w in ql for w in SPORT_KW):
        return "Спорт"
    if any(w in ql for w in ["bitcoin","ethereum","btc","eth","solana","doge","crypto"]):
        return "Крипто"
    if any(w in ql for w in ["president","election","congress","senate","democrat","republican","tariff"]):
        return "Политика"
    if any(w in ql for w in ["war","sanction","regime","missile","iran","russia","china","ukraine"]):
        return "Геополитика"
    return "Другое"

def rec_for(cat, liq):
    if cat == "Спорт" and liq >= 10_000:
        return "АНАЛИЗИРОВАТЬ", GREEN_FILL
    if cat == "Спорт" and liq >= 2_000:
        return "ОСТОРОЖНО",     YELLOW_FILL
    if cat == "Спорт":
        return "НЕТ ЛИК-ТИ",   ORANGE_FILL
    if cat == "Политика" and liq >= 5_000:
        return "LLM-АНАЛИЗ",    BLUE_FILL
    if cat in ("Геополитика","Крипто","Другое"):
        return "ПРОПУСТИТЬ",    RED_FILL
    return "ИЗУЧИТЬ", YELLOW_FILL

# Парсим активные
new_recs = []
seen_q   = set()
for m in active_raw:
    prices = m.get("outcomePrices","[]")
    if isinstance(prices,str):
        try: prices = json.loads(prices)
        except: prices = []
    if len(prices) < 2: continue
    liq  = float(m.get("liquidity",0) or 0)
    if liq < 500: continue
    yes  = float(prices[0])
    if not (0.04 < yes < 0.96): continue
    q    = m.get("question","")
    if q in seen_q: continue
    seen_q.add(q)
    end_raw = m.get("endDate","")
    end_dt = None
    if end_raw:
        try: end_dt = datetime.fromisoformat(end_raw.replace("Z","+00:00"))
        except: pass
    if end_dt and end_dt < NOW: continue
    cat = categorize(q)
    rec, rfill = rec_for(cat, liq)
    chk = (end_dt + timedelta(hours=2)) if end_dt else None
    new_recs.append({
        "q":q,"cat":cat,"yes":yes,"no":float(prices[1]),
        "liq":liq,"vol24":float(m.get("volume24hr",0) or 0),
        "end_dt":end_dt,"check_dt":chk,
        "rec":rec,"rfill":rfill,
        "pri": 0 if rec=="АНАЛИЗИРОВАТЬ" else 1 if rec=="ОСТОРОЖНО" else
               2 if rec=="LLM-АНАЛИЗ"    else 9,
    })

new_recs.sort(key=lambda x: (x["pri"], -x["liq"]))
display = new_recs[:30]
print(f"  Новых рекомендаций: {len(display)}")

HEADERS = [
    "№","Рынок","Категория","ДА","НЕТ",
    "Ликвидность $","Объём 24ч $","Дата окончания",
    "Рекомендация","Обоснование","Проверить в",
]
WHY = {
    "АНАЛИЗИРОВАТЬ": "Спорт + высокая ликв. Сравнить с Pinnacle/1xBet.",
    "ОСТОРОЖНО":     "Спорт, умеренная ликв. (<$10K). Ставить с осторожностью.",
    "НЕТ ЛИК-ТИ":   "Слишком низкая ликвидность.",
    "LLM-АНАЛИЗ":    "Политика: нужен анализ новостей. Фаза 3.",
    "ПРОПУСТИТЬ":    "Нет объективного бенчмарка вероятностей.",
    "ИЗУЧИТЬ":       "Ручной анализ. Нет чёткого бенчмарка.",
}

for r in range(2, 62):
    for c in range(1, len(HEADERS)+1):
        ws_rec.cell(row=r, column=c).value  = None
        ws_rec.cell(row=r, column=c).fill   = PatternFill()
        ws_rec.cell(row=r, column=c).border = Border()

for c, h in enumerate(HEADERS, 1):
    ws_rec.cell(row=1, column=c, value=h)
hdr(ws_rec, 1, len(HEADERS))
ws_rec.row_dimensions[1].height = 30

for i, m in enumerate(display, 1):
    r    = i + 1
    base = ZEBRA_FILL if i % 2 == 0 else PatternFill()
    ws_rec.cell(row=r, column=1,  value=i).alignment = CENTER
    ws_rec.cell(row=r, column=2,  value=m["q"]).alignment = WRAP
    ws_rec.cell(row=r, column=3,  value=m["cat"]).alignment = CENTER
    ws_rec.cell(row=r, column=4,  value=m["yes"]).number_format = "0.000"
    ws_rec.cell(row=r, column=5,  value=m["no"]).number_format  = "0.000"
    ws_rec.cell(row=r, column=6,  value=m["liq"]).number_format = '#,##0'
    ws_rec.cell(row=r, column=7,  value=m["vol24"]).number_format = '#,##0'
    if m["end_dt"]:
        ws_rec.cell(row=r, column=8, value=m["end_dt"].strftime("%Y-%m-%d %H:%M"))
    ws_rec.cell(row=r, column=9,  value=m["rec"]).alignment = CENTER
    ws_rec.cell(row=r, column=10, value=WHY.get(m["rec"],"")).alignment = WRAP
    if m["check_dt"]:
        ws_rec.cell(row=r, column=11, value=m["check_dt"].strftime("%Y-%m-%d %H:%M"))
    ws_rec.cell(row=r, column=9).fill = m["rfill"]
    for c in range(1, len(HEADERS)+1):
        ws_rec.cell(row=r, column=c).border = THIN_BORDER
        if c != 9:
            ws_rec.cell(row=r, column=c).fill = base
    ws_rec.row_dimensions[r].height = 30

col_w = {"A":4,"B":52,"C":13,"D":8,"E":8,"F":14,"G":13,"H":18,"I":18,"J":45,"K":18}
for col, w in col_w.items():
    ws_rec.column_dimensions[col].width = w

try:
    wb.save(EXCEL_PATH)
    print(f"  Excel сохранён: {EXCEL_PATH}")
except PermissionError:
    backup_path = EXCEL_PATH.parent / "polymarket_recommendations_backup.xlsx"
    wb.save(backup_path)
    print(f"  [!] Основной файл заблокирован. Сохранено в: {backup_path}")


# ════════════════════════════════════════════════════════════
# 6. ФИНАЛЬНЫЙ ОТЧЁТ В КОНСОЛЬ
# ════════════════════════════════════════════════════════════

print()
print("=" * 80)
print(f"  РЕЗУЛЬТАТЫ СТАВОК — {NOW.strftime('%d.%m.%Y %H:%M UTC')}")
print("=" * 80)

if updated:
    print()
    print("  [+] ЗАКРЫТЫЕ СТАВКИ:")
    for q, side, res, pnl, outcome in updated:
        icon = "WIN" if res == "ВЫИГРЫШ" else "LOSS"
        print(f"  [{icon}]  {side}  P&L: {pnl}")
        print(f"         {q}")
        print(f"         Исход рынка: {outcome}")
else:
    print()
    print("  Закрытых ставок с результатами не найдено через API.")
    print("  (Возможно, рынки ещё не resolved или изменился slug.)")

if still_open:
    print()
    print("  [~] АКТИВНЫЕ СТАВКИ (текущая цена / floating P&L):")
    for q, side, cur_p, fpnl, end_str in still_open:
        cpstr  = f"{cur_p:.3f}" if cur_p is not None else "?"
        fpstr  = f"${fpnl:+.2f}" if fpnl is not None else "?"
        print(f"  [ACT] {side}  Цена: {cpstr}  Float P&L: {fpstr}")
        print(f"        {q}")
        print(f"        Окончание: {end_str}")

print()
print("─" * 80)
print(f"  ИТОГО:  Закрытых={total_closed}  Побед={won_cnt}  "
      f"Проигрышей={lost_cnt}  WR={wr}%  P&L=${total_pnl:+.2f}")
print("─" * 80)

if new_recs:
    print()
    print("=" * 80)
    print("  СВЕЖИЕ РЕКОМЕНДАЦИИ (топ)")
    print("=" * 80)
    for grp_label, grp_key in [
        ("АНАЛИЗИРОВАТЬ", "АНАЛИЗИРОВАТЬ"),
        ("ОСТОРОЖНО",     "ОСТОРОЖНО"),
    ]:
        grp = [m for m in display if m["rec"] == grp_key][:6]
        if not grp: continue
        print(f"\n  -- {grp_label} --")
        for m in grp:
            end_s = m["end_dt"].strftime("%d.%m %H:%M") if m["end_dt"] else "?"
            print(f"  * {m['q'][:68]}")
            print(f"    ДА={m['yes']:.3f}  Ликв=${m['liq']:>9,.0f}  "
                  f"Конец: {end_s} UTC")
