"""
Обновление Excel: текущие цены ставок + новые рекомендации.

Пути к Excel и к JSON-кэшу рынков — через .env (см. .env.example и docs/RUNBOOK.md).
"""
from __future__ import annotations
import json, os, sys, io
from datetime import datetime, timezone, timedelta
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from dotenv import load_dotenv
from openpyxl import load_workbook
from core.bet_results import calc_pnl
from core.gamma_resolve import is_no_side, is_yes_side
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_REPO_ROOT = Path(__file__).resolve().parent
load_dotenv(_REPO_ROOT / ".env")


def _excel_path() -> Path:
    p = os.environ.get("REFRESH_EXCEL_PATH", "").strip()
    if p:
        return Path(p).expanduser().resolve()
    return _REPO_ROOT / "polymarket_рекомендации.xlsx"


def _market_cache_files() -> list[str]:
    """Пути к JSON (или .txt с JSON внутри), разделитель ; в REFRESH_MARKET_CACHE_FILES."""
    raw = os.environ.get("REFRESH_MARKET_CACHE_FILES", "").strip()
    if raw:
        return [x.strip() for x in raw.split(";") if x.strip()]
    d = os.environ.get("REFRESH_MARKET_CACHE_DIR", "").strip()
    if d:
        p = Path(d).expanduser().resolve()
        if p.is_dir():
            files = list(p.glob("*.json")) + list(p.glob("*.txt"))
            return [str(x) for x in sorted(files)]
    cache_dir = _REPO_ROOT / "data" / "market_cache"
    if cache_dir.is_dir():
        files = list(cache_dir.glob("*.json")) + list(cache_dir.glob("*.txt"))
        if files:
            return [str(x) for x in sorted(files)]
    return []


EXCEL_PATH = _excel_path()
ACTIVE_FILES = _market_cache_files()
NOW = datetime.now(timezone.utc)

# ── Стили ────────────────────────────────────────────────────
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL  = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
GREEN_FILL   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL     = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL  = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
BLUE_FILL    = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
GRAY_FILL    = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ZEBRA_FILL   = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
CENTER       = Alignment(horizontal="center", vertical="center")
WRAP         = Alignment(wrap_text=True, vertical="top")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

def hdr(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = CENTER; cell.border = THIN_BORDER

# ════════════════════════════════════════════════════════════
# 1. ЗАГРУЗКА АКТИВНЫХ РЫНКОВ
# ════════════════════════════════════════════════════════════
all_active: list[dict] = []
seen = set()
for fp in ACTIVE_FILES:
    try:
        data = json.load(open(fp, encoding="utf-8"))
    except Exception as e:
        continue
    for m in data:
        cid = m.get("conditionId","") or m.get("id","")
        if cid in seen: continue
        seen.add(cid)
        all_active.append(m)

print(f"  Активных рынков загружено: {len(all_active)}")
if not all_active:
    if not ACTIVE_FILES:
        print(
            "  [hint] Кэш рынков для «Рекомендаций» не задан: "
            "REFRESH_MARKET_CACHE_FILES / REFRESH_MARKET_CACHE_DIR или data/market_cache/*.json — см. docs/RUNBOOK.md"
        )
    else:
        print("  [hint] Указаны файлы кэша, но рынков нет — проверьте JSON в файлах.")

# Индекс по conditionId (полному и частичному) и по первым словам вопроса
idx: dict[str, dict] = {}
for m in all_active:
    cid  = m.get("conditionId","")
    slug = m.get("slug","")
    q    = (m.get("question") or "").lower().strip()
    if cid:
        idx[cid]    = m
        idx[cid[:30]] = m          # короткий вариант из Excel
    if slug:
        idx[slug] = m
    for n in (5, 4, 3):
        key = " ".join(q.split()[:n])
        if key: idx[key] = m

def get_prices(m: dict) -> tuple[float, float]:
    prices = m.get("outcomePrices","[]")
    if isinstance(prices, str):
        try: prices = json.loads(prices)
        except: prices = []
    if len(prices) >= 2:
        try: return float(prices[0]), float(prices[1])
        except: pass
    return 0.5, 0.5

def find_market(question: str, cid_short: str) -> dict | None:
    if cid_short and cid_short in idx:
        return idx[cid_short]
    q = (question or "").lower().strip()
    for n in (5, 4, 3):
        key = " ".join(q.split()[:n])
        if key in idx:
            return idx[key]
    return None

# ════════════════════════════════════════════════════════════
# 2. СТАТУС КАЖДОЙ СТАВКИ
# ════════════════════════════════════════════════════════════

# Результаты матчей 4-5 марта — известны из внешних источников
# Лига MX / Liga BetPlay (04.03.2026):
# CD Tolima vs Nacional — итог 1:1 (ничья) → Tolima НЕ победил → YES проиграл
# Club Tijuana vs Querétaro — итог 2:0 → Tijuana ПОБЕДИЛ → YES выиграл
# Puebla vs Tigres UANL — итог 0:2 → Totals 2 голов → OVER 1.5 сыграл
#   Мы ставили NO (UNDER 1.5) → ПРОИГРЫШ
# Auxerre vs Strasbourg (07.03.2026) — матч сегодня, 18:00 UTC → ещё не завершён
# Carolina Hurricanes — до 30.06.2026

MANUAL_RESULTS = {
    "AJ Auxerre vs. RC Strasbourg Alsace: O/U 1.5": {
        "outcome": "No",  # UNDER 1.5 — 0 голов
        "note":    "07.03.2026 Auxerre 0:0 Strasbourg — 0 голов. UNDER 1.5 сыграл. Мы ставили NO (UNDER) → ВЫИГРЫШ.",
    },
    "Will CD Tolima win on 2026-03-04?": {
        "outcome": "No",
        "note":    "04.03.2026 CD Tolima vs. Atletico Nacional — 1:1 (ничья). YES проиграл.",
    },
    "Club Puebla vs. Tigres de la UANL: O/U 1.5": {
        "outcome": "Yes",   # OVER сыграл
        "note":    "04.03.2026 Puebla 0:2 Tigres — 2 гола. OVER 1.5 сыграл. Мы ставили NO (UNDER) → ПРОИГРЫШ.",
    },
    "Will Club Tijuana win on 2026-03-04?": {
        "outcome": "Yes",
        "note":    "04.03.2026 Club Tijuana 2:0 Querétaro FC — победа. YES выиграл.",
    },
    "Middlesbrough FC vs. Charlton Athletic FC: O/U 2.5": {
        "outcome": "No",  # 11.03.2026 Boro 0:1 Charlton — 1 гол. UNDER 2.5. Мы YES (OVER) → ПРОИГРЫШ
        "note":    "Championship 11.03.2026. Middlesbrough 0-1 Charlton — 1 гол.",
    },
    "Kartal vs. Rybakina: Match O/U 22.5": {
        "outcome": "No",  # 11.03.2026 Rybakina 6-4 4-3 ret — 17 геймов. UNDER. Мы YES (OVER) → ПРОИГРЫШ
        "note":    "Indian Wells 11.03.2026. Rybakina def. Kartal 6-4 4-3 ret — 17 геймов.",
    },
    "Kartal vs. Rybakina: Match O/U 23.5": {
        "outcome": "No",  # 17 геймов < 23.5. UNDER. Мы YES (OVER) → ПРОИГРЫШ
        "note":    "Indian Wells 11.03.2026. 17 геймов total.",
    },
    "Game Handicap: LY (-2.5) vs LOUD (+2.5)": {
        "outcome": "No",  # Gamma: outcomes LYON/LOUD → prices [0,1] = LOUD. YES (токен LYON) проиграл
        "note":    "First Stand 17.03.2026 LYON 3-2 LOUD. Рынок: LYON только если ≥3 карт разницы; иначе LOUD. Источник: gamma-api outcomePrices.",
    },
}

wb       = load_workbook(EXCEL_PATH)
ws_bets  = wb["Ставки"]
ws_stats = wb["Статистика"]
ws_rec   = wb["Рекомендации"]

summary = []
total_pnl = 0.0
won_cnt = lost_cnt = open_cnt = 0

row = 2
while True:
    q_val = ws_bets.cell(row=row, column=3).value
    if q_val is None: break
    question  = str(q_val)
    cid_short = str(ws_bets.cell(row=row, column=2).value or "")
    side      = str(ws_bets.cell(row=row, column=4).value or "YES")
    cur_res   = ws_bets.cell(row=row, column=10).value
    try:
        bet   = float(str(ws_bets.cell(row=row, column=6).value or 10).replace("$",""))
        price = float(str(ws_bets.cell(row=row, column=7).value or 0.5))
    except:
        bet, price = 10.0, 0.5

    # Уже закрыты
    if cur_res in ("ВЫИГРЫШ", "ПРОИГРЫШ"):
        pnl_v = 0.0
        try:
            pnl_v = float(str(ws_bets.cell(row=row, column=11).value or 0))
        except: pass
        total_pnl += pnl_v
        if cur_res == "ВЫИГРЫШ": won_cnt += 1
        else:                     lost_cnt += 1
        row += 1
        continue

    end_str = str(ws_bets.cell(row=row, column=14).value or "")

    # Ручные результаты (заполненные в MANUAL_RESULTS)
    manual = MANUAL_RESULTS.get(question)
    if manual and manual.get("outcome"):
        mkt_outcome = manual["outcome"]
        if is_yes_side(side):
            won = mkt_outcome == "Yes"
        elif is_no_side(side):
            won = mkt_outcome == "No"
        else:
            row += 1
            continue
        pnl = calc_pnl(bet, price, won)

        res_ru   = "ВЫИГРЫШ" if won else "ПРОИГРЫШ"
        res_fill = GREEN_FILL if won else RED_FILL

        ws_bets.cell(row=row, column=10, value=res_ru).fill  = res_fill
        ws_bets.cell(row=row, column=11, value=round(pnl,2)).fill = res_fill
        ws_bets.cell(row=row, column=11).number_format = '#,##0.00'
        ws_bets.cell(row=row, column=12, value="Завершена (ручная)").fill = GRAY_FILL
        for c in [10,11,12]:
            ws_bets.cell(row=row, column=c).border = THIN_BORDER
            ws_bets.cell(row=row, column=c).alignment = CENTER

        pnl_str = f"+${pnl:.2f}" if pnl > 0 else f"-${abs(pnl):.2f}"
        summary.append({
            "q":      question[:60],
            "side":   side,
            "res":    res_ru,
            "pnl":    pnl_str,
            "note":   manual["note"],
        })
        total_pnl += pnl
        if won: won_cnt  += 1
        else:   lost_cnt += 1

    else:
        # Активная ставка — обновляем floating P&L из текущих цен
        open_cnt += 1
        mkt = find_market(question, cid_short)
        float_pnl = None
        cur_yes_p  = None
        if mkt:
            yes_p, no_p = get_prices(mkt)
            cur_yes_p = yes_p
            if price > 0:
                shares = bet / price
                # Для YES: стоимость = shares * yes_p. Для NO: shares * no_p
                cur_side_price = yes_p if side in ("YES", "ДА") else no_p
                float_val = shares * cur_side_price - bet
                float_pnl = round(float_val, 2)

        note = ""
        if "Auxerre" in question or "Strasbourg" in question:
            note = "Матч СЕГОДНЯ 07.03 18:00 UTC — результат будет к 20:00 UTC"
            status = "Матч сегодня!"
        elif "Carolina Hurricanes" in question:
            note = "Долгосрочная ставка (до 30.06.2026)"
            status = "Активна (долгосроч.)"
        else:
            note = "Ожидаем резолюцию Polymarket"
            status = "Ожидание API"

        if float_pnl is not None:
            ws_bets.cell(row=row, column=11, value=float_pnl)
            ws_bets.cell(row=row, column=11).number_format = '#,##0.00'
            ws_bets.cell(row=row, column=11).fill = \
                GREEN_FILL if float_pnl >= 0 else ORANGE_FILL
            ws_bets.cell(row=row, column=11).border = THIN_BORDER

        ws_bets.cell(row=row, column=12, value=status).fill = YELLOW_FILL
        ws_bets.cell(row=row, column=12).border = THIN_BORDER
        ws_bets.cell(row=row, column=12).alignment = CENTER

        summary.append({
            "q":      question[:60],
            "side":   side,
            "res":    "АКТИВНА",
            "pnl":    f"${float_pnl:+.2f}" if float_pnl is not None else "?",
            "note":   note,
        })

    row += 1

total_closed = won_cnt + lost_cnt
wr = round(won_cnt / total_closed * 100, 1) if total_closed > 0 else 0.0

# ════════════════════════════════════════════════════════════
# 3. ОБНОВЛЯЕМ СТАТИСТИКУ
# ════════════════════════════════════════════════════════════
stats = {
    2: ("Всего закрытых ставок", total_closed),
    3: ("Выигрышей",             won_cnt),
    4: ("Проигрышей",            lost_cnt),
    5: ("Win Rate (%)",          wr),
    6: ("Итоговый P&L ($)",      round(total_pnl, 2)),
    7: ("Дата обновления",       NOW.strftime("%Y-%m-%d %H:%M UTC")),
    8: ("Активных ставок",       open_cnt),
}
for rn, (label, val) in stats.items():
    ws_stats.cell(row=rn, column=1, value=label).font = Font(bold=True)
    ws_stats.cell(row=rn, column=2, value=val)
    if rn == 6:
        ws_stats.cell(row=rn, column=2).fill = \
            GREEN_FILL if total_pnl >= 0 else RED_FILL
    for c in [1,2]:
        ws_stats.cell(row=rn, column=c).border = THIN_BORDER

log_row = 12
while ws_stats.cell(row=log_row, column=1).value not in (None,""):
    log_row += 1
    if log_row > 300: break
log_msg = (f"Обновление результатов. Закрыто(ручн.): {len([s for s in summary if s['res']!='АКТИВНА'])}, "
           f"Активных: {open_cnt}. P&L=${total_pnl:+.2f}, WR={wr}%")
ws_stats.cell(row=log_row, column=1, value=NOW.strftime("%Y-%m-%d %H:%M"))
ws_stats.cell(row=log_row, column=2, value=log_msg)
ws_stats.cell(row=log_row, column=2).alignment = WRAP
for c in [1,2]:
    ws_stats.cell(row=log_row, column=c).border = THIN_BORDER
ws_stats.cell(row=log_row, column=1).fill = GREEN_FILL if total_pnl >= 0 else RED_FILL

# ════════════════════════════════════════════════════════════
# 4. ОБНОВЛЯЕМ РЕКОМЕНДАЦИИ
# ════════════════════════════════════════════════════════════
SPORT_KW = [
    "win","beat"," vs ","o/u","over","under","total","nba","nhl","nfl","ufc",
    "fc "," fc","afc","tolima","tijuana","puebla","tigres","mazatlan",
    "auxerre","strasbourg","lens","metz","gracheva","tagger","tennis",
    "hurricanes","burnley","arsenal","chelsea","liverpool","tottenham",
    "manchester","villa","lakers","celtics","warriors","nuggets","bucks",
    "fight","bout","match","toulouse","marseille","monaco","psg","lyon",
]
def categorize(q):
    ql = q.lower()
    if any(w in ql for w in SPORT_KW): return "Спорт"
    if any(w in ql for w in ["bitcoin","ethereum","btc","eth","solana","doge","crypto"]): return "Крипто"
    if any(w in ql for w in ["president","election","congress","senate","democrat","republican","tariff"]): return "Политика"
    if any(w in ql for w in ["war","sanction","iran","russia","china","ukraine","missile"]): return "Геополитика"
    return "Другое"

def rec_for(cat, liq):
    if cat == "Спорт" and liq >= 10_000: return "АНАЛИЗИРОВАТЬ", GREEN_FILL
    if cat == "Спорт" and liq >= 2_000:  return "ОСТОРОЖНО",     YELLOW_FILL
    if cat == "Спорт":                   return "НЕТ ЛИК-ТИ",   ORANGE_FILL
    if cat == "Политика" and liq >= 5000: return "LLM-АНАЛИЗ",  BLUE_FILL
    return "ПРОПУСТИТЬ", RED_FILL

WHY = {
    "АНАЛИЗИРОВАТЬ": "Спорт + высокая ликв. Сравнить с Pinnacle/1xBet.",
    "ОСТОРОЖНО":     "Спорт, умеренная ликв. (<$10K). Ставить осторожно.",
    "НЕТ ЛИК-ТИ":   "Слишком низкая ликвидность.",
    "LLM-АНАЛИЗ":    "Политика: нужен анализ новостей.",
    "ПРОПУСТИТЬ":    "Нет объективного бенчмарка.",
}
new_recs = []
seen_q = set()
for m in all_active:
    prices = m.get("outcomePrices","[]")
    if isinstance(prices,str):
        try: prices = json.loads(prices)
        except: prices = []
    if len(prices) < 2: continue
    liq  = float(m.get("liquidity",0) or 0)
    if liq < 500: continue
    yes  = float(prices[0])
    if not (0.04 < yes < 0.96): continue
    q = m.get("question","")
    if q in seen_q: continue
    seen_q.add(q)
    end_raw = m.get("endDate","")
    end_dt = None
    if end_raw:
        try: end_dt = datetime.fromisoformat(end_raw.replace("Z","+00:00"))
        except: pass
    if end_dt and end_dt < NOW: continue
    cat = categorize(q)
    rec, rfill = rec_for(cat, liq)
    chk = (end_dt + timedelta(hours=2)) if end_dt else None
    new_recs.append({
        "q":q,"cat":cat,"yes":yes,"no":float(prices[1]),
        "liq":liq,"vol24":float(m.get("volume24hr",0) or 0),
        "end_dt":end_dt,"check_dt":chk,
        "rec":rec,"rfill":rfill,
        "pri": {"АНАЛИЗИРОВАТЬ":0,"ОСТОРОЖНО":1,"LLM-АНАЛИЗ":2}.get(rec,9),
    })
new_recs.sort(key=lambda x: (x["pri"], -x["liq"]))
display = new_recs[:30]

HEADERS = ["№","Рынок","Категория","ДА","НЕТ","Ликвидность $","Объём 24ч $","Дата окончания","Рекомендация","Обоснование","Проверить в"]
for r in range(2, 62):
    for c in range(1, len(HEADERS)+1):
        ws_rec.cell(row=r, column=c).value  = None
        ws_rec.cell(row=r, column=c).fill   = PatternFill()
        ws_rec.cell(row=r, column=c).border = Border()
for c, h in enumerate(HEADERS, 1):
    ws_rec.cell(row=1, column=c, value=h)
hdr(ws_rec, 1, len(HEADERS))
ws_rec.row_dimensions[1].height = 30

for i, m in enumerate(display, 1):
    r = i + 1
    base = ZEBRA_FILL if i % 2 == 0 else PatternFill()
    ws_rec.cell(row=r, column=1,  value=i).alignment = CENTER
    ws_rec.cell(row=r, column=2,  value=m["q"]).alignment = WRAP
    ws_rec.cell(row=r, column=3,  value=m["cat"]).alignment = CENTER
    ws_rec.cell(row=r, column=4,  value=m["yes"]).number_format = "0.000"
    ws_rec.cell(row=r, column=5,  value=m["no"]).number_format  = "0.000"
    ws_rec.cell(row=r, column=6,  value=m["liq"]).number_format = '#,##0'
    ws_rec.cell(row=r, column=7,  value=m["vol24"]).number_format = '#,##0'
    if m["end_dt"]:
        ws_rec.cell(row=r, column=8, value=m["end_dt"].strftime("%Y-%m-%d %H:%M"))
    ws_rec.cell(row=r, column=9,  value=m["rec"]).alignment = CENTER
    ws_rec.cell(row=r, column=10, value=WHY.get(m["rec"],"")).alignment = WRAP
    if m["check_dt"]:
        ws_rec.cell(row=r, column=11, value=m["check_dt"].strftime("%Y-%m-%d %H:%M"))
    ws_rec.cell(row=r, column=9).fill = m["rfill"]
    for c in range(1, len(HEADERS)+1):
        ws_rec.cell(row=r, column=c).border = THIN_BORDER
        if c != 9: ws_rec.cell(row=r, column=c).fill = base
    ws_rec.row_dimensions[r].height = 30

col_w = {"A":4,"B":52,"C":13,"D":8,"E":8,"F":14,"G":13,"H":18,"I":18,"J":45,"K":18}
for col, w in col_w.items():
    ws_rec.column_dimensions[col].width = w

wb.save(EXCEL_PATH)
print(f"  Excel сохранён: {EXCEL_PATH}")

# ════════════════════════════════════════════════════════════
# 5. КОНСОЛЬНЫЙ ОТЧЁТ
# ════════════════════════════════════════════════════════════
print()
print("=" * 80)
print(f"  СТАТУС СТАВОК — {NOW.strftime('%d.%m.%Y %H:%M UTC')}")
print("=" * 80)
for s in summary:
    icon = "WIN " if s["res"]=="ВЫИГРЫШ" else "LOSS" if s["res"]=="ПРОИГРЫШ" else "ACT "
    print(f"\n  [{icon}] {s['side']}  {s['res']}  P&L: {s['pnl']}")
    print(f"  {s['q']}")
    print(f"  {s['note']}")

print()
print("─" * 80)
print(f"  Закрытых: {total_closed}  "
      f"Побед: {won_cnt}  Проигрышей: {lost_cnt}  "
      f"WR: {wr}%  Итого P&L: ${total_pnl:+.2f}")
print(f"  Активных: {open_cnt}")
print("─" * 80)

print()
print("=" * 80)
print("  СВЕЖИЕ РЕКОМЕНДАЦИИ НА СЕГОДНЯ")
print("=" * 80)
for grp_label, grp_key in [
    ("АНАЛИЗИРОВАТЬ (Спорт, высокая ликв.)", "АНАЛИЗИРОВАТЬ"),
    ("ОСТОРОЖНО (Спорт, умеренная ликв.)",   "ОСТОРОЖНО"),
]:
    grp = [m for m in display if m["rec"] == grp_key][:7]
    if not grp: continue
    print(f"\n  --- {grp_label} ---")
    for m in grp:
        end_s = m["end_dt"].strftime("%d.%m %H:%M") if m["end_dt"] else "?"
        print(f"  * {m['q'][:70]}")
        print(f"    ДА={m['yes']:.3f}  Ликв=${m['liq']:>9,.0f}  Конец: {end_s} UTC")
