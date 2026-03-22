"""Общие стили Excel и функции категоризации для скриптов обновления."""
from __future__ import annotations

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Стили ячеек
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
BLUE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ZEBRA_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

SPORT_KEYWORDS = [
    "win", "beat", " vs ", "o/u", "over", "under", "points", "score",
    "nba", "nfl", "ufc", "nhl", "mlb", "epl", "ucl", "premier", "ligue",
    "serie a", "bundesliga", "laliga", "eredivisie", "championship",
    " fc", "fc ", " sc", "sc ", "cf ", "afc",
    "united", " city", "real ", "atletico", "barcelona", "juventus",
    "bayern", "dortmund", "psg", "arsenal", "chelsea", "liverpool",
    "tottenham", "manchester", "inter ", "milan ", "napoli", "roma ",
    "burnley", "villa", "wolves", "brentford", "everton", "brighton",
    "tolima", "tijuana", "puebla", "tigres", "auxerre", "strasbourg",
    "lens", "metz", "gracheva", "tagger", "tennis", "hurricanes",
    "lakers", "celtics", "heat", "warriors", "bulls", "nets", "knicks",
    "mazatlan", "fight", "bout", "match", "toulouse", "marseille", "monaco", "lyon",
]


def excel_header(ws, row: int, cols: int) -> None:
    """Оформить заголовок таблицы."""
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def categorize_market(question: str) -> str:
    """Категория рынка по вопросу: Спорт, Крипто, Политика, Геополитика, Другое."""
    ql = (question or "").lower()
    if any(w in ql for w in SPORT_KEYWORDS):
        return "Спорт"
    if any(w in ql for w in ["bitcoin", "ethereum", "btc", "eth", "solana", "doge", "crypto"]):
        return "Крипто"
    if any(w in ql for w in ["president", "election", "congress", "senate", "democrat", "republican", "tariff"]):
        return "Политика"
    if any(w in ql for w in ["war", "sanction", "iran", "russia", "china", "ukraine", "missile"]):
        return "Геополитика"
    return "Другое"
