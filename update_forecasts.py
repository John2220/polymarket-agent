"""
Обновление прогнозов и результатов в Excel.
Загружает свежие данные с Polymarket API, обновляет вкладки Рекомендации и Ставки.
"""
from __future__ import annotations
import json, sys, io, requests
from datetime import datetime, timezone, timedelta
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from core.bet_results import calc_pnl
from core.gamma_resolve import bet_won_for_binary_market, market_fully_resolved

GAMMA_URL = "https://gamma-api.polymarket.com"
EXCEL_PATH = Path(__file__).parent / "polymarket_рекомендации.xlsx"
CACHE_DIR = Path(__file__).parent / "cache"
NOW = datetime.now(timezone.utc)
TODAY = NOW.strftime("%Y-%m-%d")

CACHE_DIR.mkdir(exist_ok=True)

# ── Стили ────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
BLUE_FILL   = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
GRAY_FILL   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ZEBRA_FILL  = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
CENTER      = Alignment(horizontal="center", vertical="center")
WRAP        = Alignment(wrap_text=True, vertical="top")
THIN        = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

def hdr(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = HEADER_FILL
        cell.alignment = CENTER; cell.border = THIN


def fetch_markets(url_suffix: str, params: dict) -> list:
    """Загрузить рынки с Gamma API."""
    url = f"{GAMMA_URL}{url_suffix}"
    try:
        r = requests.get(url, params=params, timeout=15)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        print(f"  [!] Ошибка загрузки {url}: {e}")
        return []


SPORT_KW = [
    "win","beat"," vs ","o/u","over","under","nba","nhl","nfl","ufc",
    "fc "," fc","afc","tolima","tijuana","puebla","tigres","auxerre",
    "strasbourg","lens","metz","gracheva","tagger","tennis","hurricanes",
    "burnley","arsenal","chelsea","liverpool","tottenham","manchester",
]

def categorize(q: str) -> str:
    ql = q.lower()
    if any(w in ql for w in SPORT_KW): return "Спорт"
    if any(w in ql for w in ["bitcoin","ethereum","btc","eth","solana","crypto"]): return "Крипто"
    if any(w in ql for w in ["president","election","congress","democrat","republican"]): return "Политика"
    if any(w in ql for w in ["war","sanction","iran","russia","ukraine"]): return "Геополитика"
    return "Другое"

def rec_for(cat: str, liq: float) -> tuple[str, PatternFill]:
    if cat == "Спорт" and liq >= 10_000: return "АНАЛИЗИРОВАТЬ", GREEN_FILL
    if cat == "Спорт" and liq >= 2_000:  return "ОСТОРОЖНО", YELLOW_FILL
    if cat == "Спорт":                   return "НЕТ ЛИК-ТИ", ORANGE_FILL
    if cat == "Политика" and liq >= 5_000: return "LLM-АНАЛИЗ", BLUE_FILL
    return "ПРОПУСТИТЬ", RED_FILL

WHY = {"АНАЛИЗИРОВАТЬ":"Спорт + высокая ликв. Сравнить с Pinnacle.",
       "ОСТОРОЖНО":"Спорт, умеренная ликв.","НЕТ ЛИК-ТИ":"Низкая ликвидность.",
       "LLM-АНАЛИЗ":"Политика: анализ новостей.","ПРОПУСТИТЬ":"Нет бенчмарка."}


def main():
    print(f"  Обновление прогнозов — {TODAY} {NOW.strftime('%H:%M')} UTC")
    print()

    # 1. Загрузка активных рынков
    active = fetch_markets("/markets", {
        "active": "true", "closed": "false", "limit": 100,
        "order": "liquidity", "ascending": "false"
    })
    soccer = fetch_markets("/markets", {
        "active": "true", "closed": "false", "limit": 50, "tag": "soccer",
        "order": "liquidity", "ascending": "false"
    })
    closed = fetch_markets("/markets", {
        "closed": "true", "limit": 30, "order": "endDate", "ascending": "false"
    })

    all_markets: list[dict] = []
    seen = set()
    for data in [active, soccer]:
        for m in data:
            cid = m.get("conditionId","") or m.get("id","")
            if cid in seen: continue
            seen.add(cid)
            all_markets.append(m)
    print(f"  Активных рынков: {len(all_markets)}")
    print(f"  Закрытых: {len(closed)}")

    # 2. Парсинг и сортировка
    parsed = []
    for m in all_markets:
        prices = m.get("outcomePrices","[]")
        if isinstance(prices, str):
            try: prices = json.loads(prices)
            except: prices = []
        if len(prices) < 2: continue
        liq = float(m.get("liquidity",0) or 0)
        if liq < 500: continue
        yes = float(prices[0])
        if not (0.04 < yes < 0.96): continue
        end_raw = m.get("endDate","")
        end_dt = None
        if end_raw:
            try: end_dt = datetime.fromisoformat(end_raw.replace("Z","+00:00"))
            except: pass
        if end_dt and end_dt < NOW: continue
        cat = categorize(m.get("question",""))
        rec, rfill = rec_for(cat, liq)
        pri = 0 if rec=="АНАЛИЗИРОВАТЬ" else 1 if rec=="ОСТОРОЖНО" else 9
        parsed.append({
            "q": m.get("question",""), "cat": cat, "yes": yes, "no": float(prices[1]),
            "liq": liq, "vol24": float(m.get("volume24hr",0) or 0),
            "end_dt": end_dt, "check_dt": (end_dt + timedelta(hours=2)) if end_dt else None,
            "rec": rec, "rfill": rfill, "pri": pri,
        })
    parsed.sort(key=lambda x: (x["pri"], -x["liq"]))
    display = parsed[:30]
    print(f"  Рекомендаций: {len(display)}")

    # 3. Индекс закрытых для проверки результатов
    closed_idx = {}
    for m in closed:
        q = (m.get("question") or "").lower()
        cid = m.get("conditionId","")
        if cid: closed_idx[cid] = m
        for n in (5,4,3):
            key = " ".join(q.split()[:n])
            if key: closed_idx[key] = m

    # 4. Открываем Excel
    wb = load_workbook(EXCEL_PATH)
    ws_rec = wb["Рекомендации"]
    ws_bets = wb["Ставки"]
    ws_stats = wb["Статистика"]

    # 5. Обновляем ставки — ищем результаты в закрытых
    total_pnl = 0.0
    won, lost = 0, 0
    row = 2
    while True:
        q = ws_bets.cell(row=row, column=3).value
        if q is None: break
        res = ws_bets.cell(row=row, column=10).value
        if res in ("ВЫИГРЫШ","ПРОИГРЫШ"):
            try: total_pnl += float(ws_bets.cell(row=row, column=11).value or 0)
            except: pass
            if res == "ВЫИГРЫШ": won += 1
            else: lost += 1
            row += 1
            continue
        cid = str(ws_bets.cell(row=row, column=2).value or "")
        side = str(ws_bets.cell(row=row, column=4).value or "YES")
        try:
            bet = float(str(ws_bets.cell(row=row, column=6).value or 10))
            price = float(str(ws_bets.cell(row=row, column=7).value or 0.5))
        except: bet, price = 10.0, 0.5

        found = None
        ql = (q or "").lower()
        for n in (5,4,3):
            key = " ".join(ql.split()[:n])
            if key in closed_idx:
                m = closed_idx[key]
                if market_fully_resolved(m):
                    found = m
                    break
        if found:
            won_bet = bet_won_for_binary_market(side, found)
            if won_bet is None:
                row += 1
                continue
            pnl = calc_pnl(bet, price, won_bet)
            res_ru = "ВЫИГРЫШ" if won_bet else "ПРОИГРЫШ"
            fill = GREEN_FILL if won_bet else RED_FILL
            ws_bets.cell(row=row, column=10, value=res_ru).fill = fill
            ws_bets.cell(row=row, column=11, value=round(pnl,2)).fill = fill
            ws_bets.cell(row=row, column=11).number_format = '#,##0.00'
            ws_bets.cell(row=row, column=12, value="Завершена").fill = GRAY_FILL
            total_pnl += pnl
            if won_bet: won += 1
            else: lost += 1
            print(f"  [Обновлено] {res_ru} {q[:50]} P&L=${pnl:+.2f}")
        row += 1

    total_closed = won + lost
    wr = round(won / total_closed * 100, 1) if total_closed > 0 else 0

    # 6. Обновляем Рекомендации
    HEADERS = ["№","Рынок","Категория","ДА","НЕТ","Ликвидность $","Объём 24ч $","Дата окончания","Рекомендация","Обоснование","Проверить в"]
    for r in range(2, 62):
        for c in range(1, 12):
            ws_rec.cell(row=r, column=c).value = None
            ws_rec.cell(row=r, column=c).fill = PatternFill()
    for c, h in enumerate(HEADERS, 1):
        ws_rec.cell(row=1, column=c, value=h)
    hdr(ws_rec, 1, len(HEADERS))
    for i, m in enumerate(display, 1):
        r = i + 1
        base = ZEBRA_FILL if i % 2 == 0 else PatternFill()
        ws_rec.cell(row=r, column=1, value=i).alignment = CENTER
        ws_rec.cell(row=r, column=2, value=m["q"]).alignment = WRAP
        ws_rec.cell(row=r, column=3, value=m["cat"]).alignment = CENTER
        ws_rec.cell(row=r, column=4, value=m["yes"]).number_format = "0.000"
        ws_rec.cell(row=r, column=5, value=m["no"]).number_format = "0.000"
        ws_rec.cell(row=r, column=6, value=m["liq"]).number_format = '#,##0'
        ws_rec.cell(row=r, column=7, value=m["vol24"]).number_format = '#,##0'
        if m["end_dt"]:
            ws_rec.cell(row=r, column=8, value=m["end_dt"].strftime("%Y-%m-%d %H:%M"))
        ws_rec.cell(row=r, column=9, value=m["rec"]).alignment = CENTER
        ws_rec.cell(row=r, column=9).fill = m["rfill"]
        ws_rec.cell(row=r, column=10, value=WHY.get(m["rec"],"")).alignment = WRAP
        if m["check_dt"]:
            ws_rec.cell(row=r, column=11, value=m["check_dt"].strftime("%Y-%m-%d %H:%M"))
        for c in range(1, 12):
            ws_rec.cell(row=r, column=c).border = THIN
            if c != 9: ws_rec.cell(row=r, column=c).fill = base

    # 7. Статистика
    for rn, (label, val) in [
        (2, ("Закрытых ставок", total_closed)),
        (3, ("Выигрышей", won)),
        (4, ("Проигрышей", lost)),
        (5, ("Win Rate (%)", wr)),
        (6, ("Итоговый P&L ($)", round(total_pnl, 2))),
        (7, ("Обновлено", NOW.strftime("%Y-%m-%d %H:%M UTC"))),
    ]:
        ws_stats.cell(row=rn, column=1, value=label).font = Font(bold=True)
        ws_stats.cell(row=rn, column=2, value=val)
        if rn == 6: ws_stats.cell(row=rn, column=2).fill = GREEN_FILL if total_pnl >= 0 else RED_FILL
        for c in [1,2]: ws_stats.cell(row=rn, column=c).border = THIN

    # Журнал
    log_row = 12
    while ws_stats.cell(row=log_row, column=1).value:
        log_row += 1
        if log_row > 200: break
    ws_stats.cell(row=log_row, column=1, value=NOW.strftime("%Y-%m-%d %H:%M"))
    ws_stats.cell(row=log_row, column=2,
        value=f"update_forecasts: {len(display)} рекомендаций, P&L=${total_pnl:+.2f}, WR={wr}%")
    ws_stats.cell(row=log_row, column=2).alignment = WRAP
    for c in [1,2]: ws_stats.cell(row=log_row, column=c).border = THIN

    wb.save(EXCEL_PATH)
    print()
    print(f"  Excel сохранён: {EXCEL_PATH}")
    print(f"  Ставок: {total_closed}  Побед: {won}  Проигрышей: {lost}  P&L: ${total_pnl:+.2f}  WR: {wr}%")
    print()
    print("  Топ рекомендаций:")
    for m in [x for x in display if x["rec"] in ("АНАЛИЗИРОВАТЬ","ОСТОРОЖНО")][:8]:
        end_s = m["end_dt"].strftime("%d.%m %H:%M") if m["end_dt"] else "?"
        print(f"    [{m['rec']}] {m['q'][:55]}")
        print(f"       ДА={m['yes']:.3f}  Ликв=${m['liq']:,.0f}  Конец: {end_s}")


if __name__ == "__main__":
    main()
