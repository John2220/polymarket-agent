"""
Обновление результатов и новых рекомендаций в Excel.
"""
from __future__ import annotations
import json, sys, io
from datetime import datetime, timezone, timedelta
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.bet_results import calc_pnl
from core.gamma_resolve import bet_won_for_binary_market, market_fully_resolved

# ── Конфигурация ─────────────────────────────────────────────
FILES = {
    "liquidity": r"C:\Users\Lomov\.cursor\projects\c-Users-Lomov-Documents-Arduino-libraries-M5Unified-src-utility\agent-tools\a7a7e228-f77e-4274-a05c-93be9c257d0c.txt",
    "soccer":    r"C:\Users\Lomov\.cursor\projects\c-Users-Lomov-Documents-Arduino-libraries-M5Unified-src-utility\agent-tools\b1e09a7a-6849-4a48-80cf-bf37ca3ca89a.txt",
    "nba":       r"C:\Users\Lomov\.cursor\projects\c-Users-Lomov-Documents-Arduino-libraries-M5Unified-src-utility\agent-tools\705d3e9b-be1f-4ae6-9215-2791ae7bb6d9.txt",
    "closed":    r"C:\Users\Lomov\.cursor\projects\c-Users-Lomov-Documents-Arduino-libraries-M5Unified-src-utility\agent-tools\f9997b84-0bf5-414f-a66c-707c5e0fe29a.txt",
}
EXCEL_PATH = Path(r"C:\Users\Lomov\Desktop\polymarket-agent\polymarket_рекомендации.xlsx")
NOW = datetime.now(timezone.utc)
TODAY = NOW.strftime("%Y-%m-%d")

# ── Стили ────────────────────────────────────────────────────
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL  = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
GREEN_FILL   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL     = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL  = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
BLUE_FILL    = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
GRAY_FILL    = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ZEBRA_FILL   = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
WRAP         = Alignment(wrap_text=True, vertical="top")
CENTER       = Alignment(horizontal="center", vertical="center")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

def hdr(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = CENTER; cell.border = THIN_BORDER

# ════════════════════════════════════════════════════════════
# 1. ЗАГРУЗКА ДАННЫХ
# ════════════════════════════════════════════════════════════

def load_json(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"  [!] Ошибка чтения {path}: {e}")
        return []

# Загрузка и дедупликация
all_raw: list[dict] = []
seen = set()
total_by_src = {}
for key, fpath in FILES.items():
    if key == "closed":
        continue
    data = load_json(fpath)
    total_by_src[key] = len(data)
    for m in data:
        cid = m.get("conditionId") or m.get("id", "")
        if cid in seen:
            continue
        seen.add(cid)
        all_raw.append(m)

closed_raw = load_json(FILES["closed"])
print(f"  Загружено рынков: {sum(total_by_src.values())} (уникальных: {len(all_raw)})")
print(f"  Закрытых рынков для проверки результатов: {len(closed_raw)}")


# ════════════════════════════════════════════════════════════
# 2. ПАРСИНГ РЫНКОВ
# ════════════════════════════════════════════════════════════

SPORT_KEYWORDS = [
    "win", "beat", "defeat", "o/u", "over", "under", "points", "score",
    "nba", "nfl", "ufc", "nhl", "mlb", "epl", "ucl", "premier", "ligue",
    "serie a", "bundesliga", "laliga", "eredivisie", "championship",
    " fc", "fc ", " sc", "sc ", "cf ", "afc",
    "united", " city", "real ", "atletico", "barcelona", "juventus",
    "bayern", "dortmund", "psg", "arsenal", "chelsea", "liverpool",
    "tottenham", "manchester", "inter ", "milan ", "napoli", "roma ",
    "burnley", "villa", "wolves", "brentford", "everton", "brighton",
    "newcastle", "leicester", "ipswich", "southampton", "crystal",
    "fulham", "nottingham", "bournemouth", "lakers", "celtics", "heat",
    "warriors", "bulls", "nets", "knicks", "76ers", "raptors", "bucks",
    "nuggets", "clippers", "suns", "mavericks", "spurs", "rockets",
    "thunder", "blazers", "hawks", "hornets", "magic", "cavaliers",
    "pistons", "pacers", "grizzlies", "jazz", "timberwolves", "pelicans",
    "kings", "tennis", "formula 1", "f1", "golf", "cricket",
    "fight", "bout", "match", " vs ", " vs.", "game",
]

def categorize(q: str) -> str:
    ql = q.lower()
    if any(w in ql for w in SPORT_KEYWORDS):
        return "Спорт"
    if any(w in ql for w in ["bitcoin","ethereum","btc","eth","solana","doge","crypto","token"]):
        return "Крипто"
    if any(w in ql for w in ["president","election","congress","senate","democrat","republican","gdp","fed rate","tariff"]):
        return "Политика"
    if any(w in ql for w in ["war","sanction","regime","missile","attack","iran","russia","china","ukraine"]):
        return "Геополитика"
    if any(w in ql for w in ["temperature","weather","rain","snow","hurricane","earthquake"]):
        return "Погода"
    if any(w in ql for w in ["alien","ufo","jesus","return of","bigfoot","flat earth"]):
        return "Экзотика"
    return "Другое"

def parse_market(m: dict) -> dict | None:
    prices = m.get("outcomePrices", "[]")
    if isinstance(prices, str):
        try:
            prices = json.loads(prices)
        except Exception:
            prices = []
    liq   = float(m.get("liquidity",  0) or 0)
    vol   = float(m.get("volume",     0) or 0)
    vol24 = float(m.get("volume24hr", 0) or 0)
    if liq < 100 or len(prices) < 2:
        return None
    yes_p = float(prices[0])
    no_p  = float(prices[1])
    if not (0.02 < yes_p < 0.98):
        return None

    q   = m.get("question", "")
    cat = categorize(q)

    end_raw = m.get("endDate", "")
    end_dt = check_dt = None
    if end_raw:
        try:
            end_dt   = datetime.fromisoformat(end_raw.replace("Z", "+00:00"))
            check_dt = end_dt + timedelta(hours=2)
        except Exception:
            pass

    # Уже закончился?
    expired = end_dt and end_dt < NOW

    # Рекомендация
    if cat == "Спорт" and liq >= 10_000:
        rec, why = "АНАЛИЗИРОВАТЬ", "Спорт + высокая ликв. Сравнить с Pinnacle/1xBet."
        rfill = GREEN_FILL
    elif cat == "Спорт" and liq >= 2_000:
        rec, why = "ОСТОРОЖНО", "Спорт, но ликвидность умеренная (<$10K)."
        rfill = YELLOW_FILL
    elif cat == "Спорт":
        rec, why = "НЕТ ЛИК-ТИ", "Спорт, но ликвидность мала (<$2K)."
        rfill = ORANGE_FILL
    elif cat == "Политика" and liq >= 5_000:
        rec, why = "LLM-АНАЛИЗ", "Политика: нужен анализ новостей. Высокая ликв."
        rfill = BLUE_FILL
    elif cat in ("Геополитика", "Погода", "Экзотика"):
        rec, why = "ПРОПУСТИТЬ", "Нет объективного бенчмарка вероятностей."
        rfill = RED_FILL
    elif cat == "Крипто":
        rec, why = "ПРОПУСТИТЬ", "Крипто: эффективный рынок."
        rfill = RED_FILL
    else:
        rec, why = "ИЗУЧИТЬ", "Нет чёткого бенчмарка. Ручной анализ."
        rfill = YELLOW_FILL

    if expired:
        rec = "ИСТЁК"
        rfill = GRAY_FILL

    # Приоритет (0=лучший)
    priority = 0
    if rec == "АНАЛИЗИРОВАТЬ":
        priority = 0
    elif rec == "ОСТОРОЖНО":
        priority = 1
    elif rec == "LLM-АНАЛИЗ":
        priority = 2
    elif rec == "ИЗУЧИТЬ":
        priority = 3
    else:
        priority = 9

    return {
        "question": q, "cat": cat, "yes": yes_p, "no": no_p,
        "liq": liq, "vol": vol, "vol24": vol24,
        "end_dt": end_dt, "check_dt": check_dt,
        "rec": rec, "why": why, "rfill": rfill,
        "priority": priority, "expired": expired,
        "slug": m.get("slug",""), "cid": m.get("conditionId",""),
    }

parsed = [p for m in all_raw if (p := parse_market(m)) is not None and not p["expired"]]
parsed.sort(key=lambda x: (x["priority"], -x["liq"]))
print(f"  Действующих рынков после фильтрации: {len(parsed)}")

sport_top = [p for p in parsed if p["cat"] == "Спорт" and p["rec"] in ("АНАЛИЗИРОВАТЬ","ОСТОРОЖНО")]
print(f"  Спортивных для ставок: {len(sport_top)}")


# ════════════════════════════════════════════════════════════
# 3. ПРОВЕРКА РЕЗУЛЬТАТОВ ИЗ ЗАКРЫТЫХ РЫНКОВ
# ════════════════════════════════════════════════════════════

# Строим словарь закрытых рынков
closed_map: dict[str, dict] = {}
for m in closed_raw:
    q = m.get("question","").lower()
    cid = m.get("conditionId","")
    if cid:
        closed_map[cid] = m
    # Также индексируем по первым словам вопроса для fuzzy match
    if q:
        key = " ".join(q.split()[:4])
        closed_map[key] = m

def find_result(question: str, cid: str = "") -> dict | None:
    """Возвращает сырой объект рынка, если закрыт с определённым Yes/No-исходом."""
    m = None
    if cid and cid in closed_map:
        m = closed_map[cid]
    if m is None:
        q = question.lower()
        key = " ".join(q.split()[:4])
        if key in closed_map:
            m = closed_map[key]
    if m is not None and market_fully_resolved(m):
        return m
    return None


# ════════════════════════════════════════════════════════════
# 4. ОТКРЫВАЕМ EXCEL И ОБНОВЛЯЕМ
# ════════════════════════════════════════════════════════════

wb = load_workbook(EXCEL_PATH)
ws_rec   = wb["Рекомендации"]
ws_bets  = wb["Ставки"]
ws_stats = wb["Статистика"]

# ── 4а. Обновляем вкладку СТАВКИ ─────────────────────────────
updated_bets = []
row = 2
while True:
    q_cell  = ws_bets.cell(row=row, column=3)
    res_cell = ws_bets.cell(row=row, column=10)
    if q_cell.value is None:
        break
    question = str(q_cell.value)
    cid      = str(ws_bets.cell(row=row, column=2).value or "")
    cur_res  = res_cell.value

    if cur_res not in (None, "", "Ожидание"):
        row += 1
        continue  # уже есть результат

    # Ищем в закрытых рынках
    closed_m = find_result(question, cid)
    if closed_m:
        side = str(ws_bets.cell(row=row, column=4).value or "YES")
        bet_price_cell = ws_bets.cell(row=row, column=7).value or 0.5
        bet_usd_cell   = ws_bets.cell(row=row, column=6).value or 10
        try:
            price = float(bet_price_cell)
            bet   = float(str(bet_usd_cell).replace("$",""))
        except Exception:
            price, bet = 0.5, 10.0

        won = bet_won_for_binary_market(side, closed_m)
        if won is None:
            row += 1
            continue
        pnl = calc_pnl(bet, price, won)
        res_ru = "ВЫИГРЫШ" if won else "ПРОИГРЫШ"

        ws_bets.cell(row=row, column=10, value=res_ru).fill = GREEN_FILL if won else RED_FILL
        ws_bets.cell(row=row, column=11, value=pnl).fill    = GREEN_FILL if won else RED_FILL
        ws_bets.cell(row=row, column=11).number_format = '#,##0.00'
        ws_bets.cell(row=row, column=12, value="Завершена").fill = GRAY_FILL
        updated_bets.append((question[:50], res_ru, pnl))
        print(f"  Ставка обновлена: [{res_ru}] {question[:55]}")
    row += 1

if not updated_bets:
    print("  Новых результатов через API не найдено.")

# ── 4б. Вкладка РЕКОМЕНДАЦИИ: обновляем ─────────────────────

HEADERS = [
    "№","Рынок","Категория","ДА","НЕТ",
    "Ликвидность $","Объём 24ч $","Дата окончания",
    "Рекомендация","Обоснование","Проверить в",
]

# Очищаем строки 2..60
for r in range(2, 61):
    for c in range(1, len(HEADERS)+1):
        cell = ws_rec.cell(row=r, column=c)
        cell.value = None
        cell.fill  = PatternFill()
        cell.border = Border()

for c, h in enumerate(HEADERS, 1):
    ws_rec.cell(row=1, column=c, value=h)
hdr(ws_rec, 1, len(HEADERS))
ws_rec.row_dimensions[1].height = 30

# Берём топ-30 рынков
display = parsed[:30]
for i, m in enumerate(display, 1):
    r = i + 1
    base = ZEBRA_FILL if i % 2 == 0 else PatternFill()

    ws_rec.cell(row=r, column=1,  value=i).alignment = CENTER
    ws_rec.cell(row=r, column=2,  value=m["question"]).alignment = WRAP
    ws_rec.cell(row=r, column=3,  value=m["cat"]).alignment = CENTER
    ws_rec.cell(row=r, column=4,  value=m["yes"]).number_format = "0.000"
    ws_rec.cell(row=r, column=5,  value=m["no"]).number_format  = "0.000"
    ws_rec.cell(row=r, column=6,  value=m["liq"]).number_format = '#,##0'
    ws_rec.cell(row=r, column=7,  value=m["vol24"]).number_format = '#,##0'
    if m["end_dt"]:
        ws_rec.cell(row=r, column=8,  value=m["end_dt"].strftime("%Y-%m-%d %H:%M"))
    ws_rec.cell(row=r, column=9,  value=m["rec"]).alignment = CENTER
    ws_rec.cell(row=r, column=10, value=m["why"]).alignment = WRAP
    if m["check_dt"]:
        ws_rec.cell(row=r, column=11, value=m["check_dt"].strftime("%Y-%m-%d %H:%M"))

    ws_rec.cell(row=r, column=9).fill = m["rfill"]
    for c in range(1, len(HEADERS)+1):
        cell = ws_rec.cell(row=r, column=c)
        cell.border = THIN_BORDER
        if c != 9:
            cell.fill = base
    ws_rec.row_dimensions[r].height = 30

# Ширина столбцов
col_widths = {"A":4,"B":52,"C":13,"D":8,"E":8,"F":14,"G":13,"H":18,"I":18,"J":45,"K":18}
for col, w in col_widths.items():
    ws_rec.column_dimensions[col].width = w

# ── 4в. Статистика: итоговая строка ──────────────────────────

# Считаем P&L по всем ставкам
total_bets = won_count = lost_count = 0
total_pnl = 0.0
r = 2
while True:
    q = ws_bets.cell(row=r, column=3).value
    if q is None:
        break
    res = ws_bets.cell(row=r, column=10).value or ""
    pnl_val = ws_bets.cell(row=r, column=11).value or 0
    try:
        pnl_val = float(str(pnl_val).replace("$","").replace("+",""))
    except Exception:
        pnl_val = 0.0
    if res in ("ВЫИГРЫШ","ПРОИГРЫШ"):
        total_bets += 1
        total_pnl  += pnl_val
        if res == "ВЫИГРЫШ":
            won_count += 1
        else:
            lost_count += 1
    r += 1

wr = round(won_count / total_bets * 100, 1) if total_bets > 0 else 0

# Обновляем ячейки статистики (строки 2-10)
stats_data = {
    2: ("Всего ставок (закрытых)", total_bets),
    3: ("Выигрышей",               won_count),
    4: ("Проигрышей",              lost_count),
    5: ("Процент побед (%)",       wr),
    6: ("Итоговый P&L ($)",        round(total_pnl, 2)),
    7: ("Дата последнего обновл.", NOW.strftime("%Y-%m-%d %H:%M UTC")),
    8: ("Рынков проанализировано", len(parsed)),
    9: ("Спортивных к торговле",   len(sport_top)),
}
for row_n, (label, val) in stats_data.items():
    ws_stats.cell(row=row_n, column=1, value=label).font = Font(bold=True)
    ws_stats.cell(row=row_n, column=2, value=val)
    ws_stats.cell(row=row_n, column=1).border = THIN_BORDER
    ws_stats.cell(row=row_n, column=2).border = THIN_BORDER

# Журнал событий
log_row = 12
while ws_stats.cell(row=log_row, column=1).value not in (None, ""):
    log_row += 1
    if log_row > 200:
        break

upd_summary = f"Новых результатов: {len(updated_bets)}, новых рекомендаций: {len(display)}, P&L={round(total_pnl,2)}"
ws_stats.cell(row=log_row, column=1, value=NOW.strftime("%Y-%m-%d %H:%M")).border = THIN_BORDER
ws_stats.cell(row=log_row, column=2, value=upd_summary).border = THIN_BORDER
ws_stats.cell(row=log_row, column=2).alignment = WRAP
ws_stats.cell(row=log_row, column=1).fill = BLUE_FILL

wb.save(EXCEL_PATH)
print(f"\n  Excel сохранён: {EXCEL_PATH}")


# ════════════════════════════════════════════════════════════
# 5. ВЫВОД ТОПА РЕКОМЕНДАЦИЙ В КОНСОЛЬ
# ════════════════════════════════════════════════════════════

print()
print("=" * 80)
print(f"  ТОП РЕКОМЕНДАЦИИ НА {TODAY}")
print("=" * 80)

sections = [
    ("АНАЛИЗИРОВАТЬ (Спорт, высокая ликвидность)", "АНАЛИЗИРОВАТЬ"),
    ("ОСТОРОЖНО (Спорт, умеренная ликвидность)",   "ОСТОРОЖНО"),
    ("LLM-АНАЛИЗ (Политика)",                       "LLM-АНАЛИЗ"),
]
for title, rec_key in sections:
    group = [m for m in display if m["rec"] == rec_key]
    if not group:
        continue
    print(f"\n  --- {title} ---")
    for m in group[:8]:
        end_str = m["end_dt"].strftime("%d.%m.%Y %H:%M UTC") if m["end_dt"] else "?"
        print(f"  * {m['question'][:65]}")
        print(f"    ДА={m['yes']:.3f}  НЕТ={m['no']:.3f}  "
              f"Ликв.=${m['liq']:>10,.0f}  Конец: {end_str}")

print()
print("=" * 80)
print(f"  Итоги:  Ставок={total_bets}  Побед={won_count}  "
      f"Проигрышей={lost_count}  P&L=${total_pnl:+.2f}  WR={wr}%")
print("=" * 80)
