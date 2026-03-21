"""
Выбор рынков, расчёт ставок по Kelly, запись в Excel.
Режим: рекомендательный (бумажные ставки — без реального размещения ордеров).

⚠️ ВНИМАНИЕ: Этот скрипт НЕ использует Pinnacle. Edge оценивается приближённо (EDGE_EST).
Для торговли с реальным edge от sharp-букмекера используйте: python main.py --mode recommend
"""
from __future__ import annotations
import json, sys, io
from datetime import datetime, timezone, timedelta
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Параметры ─────────────────────────────────────────────────
BANKROLL    = 1_000.0   # Bankroll в USD
KELLY_FRAC  = 0.25      # Дробный Kelly (четверть)
MAX_BET_USD = 30.0      # Максимум на одну ставку
MIN_BET_USD = 5.0       # Минимум на одну ставку
# Предполагаемый edge (оценка — без Pinnacle). Для реального edge — main.py --mode recommend
EDGE_EST    = 0.05

EXCEL_PATH = Path(r"C:\Users\Lomov\Desktop\polymarket-agent\polymarket_рекомендации.xlsx")
NOW = datetime.now(timezone.utc)

# ── Стили ────────────────────────────────────────────────────
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL  = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
GREEN_FILL   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL     = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BLUE_FILL    = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
ORANGE_FILL  = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
GRAY_FILL    = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ZEBRA_FILL   = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
CENTER       = Alignment(horizontal="center", vertical="center")
WRAP         = Alignment(wrap_text=True, vertical="top")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

def hdr(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = CENTER; cell.border = THIN_BORDER

# ════════════════════════════════════════════════════════════
# 1. ЗАГРУЗКА СВЕЖИХ РЫНКОВ
# ════════════════════════════════════════════════════════════

FILES = [
    r"C:\Users\Lomov\.cursor\projects\c-Users-Lomov-Documents-Arduino-libraries-M5Unified-src-utility\agent-tools\a7a7e228-f77e-4274-a05c-93be9c257d0c.txt",
    r"C:\Users\Lomov\.cursor\projects\c-Users-Lomov-Documents-Arduino-libraries-M5Unified-src-utility\agent-tools\b1e09a7a-6849-4a48-80cf-bf37ca3ca89a.txt",
    r"C:\Users\Lomov\.cursor\projects\c-Users-Lomov-Documents-Arduino-libraries-M5Unified-src-utility\agent-tools\705d3e9b-be1f-4ae6-9215-2791ae7bb6d9.txt",
]

seen_cids = set()
all_markets: list[dict] = []

for fp in FILES:
    try:
        data = json.load(open(fp, encoding="utf-8"))
    except Exception as e:
        print(f"  [!] Ошибка чтения {fp}: {e}")
        continue
    for m in data:
        cid = m.get("conditionId", "") or m.get("id", "")
        if cid in seen_cids:
            continue
        seen_cids.add(cid)
        prices = m.get("outcomePrices", "[]")
        if isinstance(prices, str):
            try:
                prices = json.loads(prices)
            except Exception:
                prices = []
        if len(prices) < 2:
            continue
        liq   = float(m.get("liquidity",  0) or 0)
        vol24 = float(m.get("volume24hr", 0) or 0)
        if liq < 2_000:
            continue
        yes_p = float(prices[0])
        no_p  = float(prices[1])
        if not (0.05 < yes_p < 0.95):
            continue
        end_raw = m.get("endDate", "")
        end_dt = None
        if end_raw:
            try:
                end_dt = datetime.fromisoformat(end_raw.replace("Z", "+00:00"))
            except Exception:
                pass
        # Пропускаем уже истёкшие
        if end_dt and end_dt < NOW:
            continue
        all_markets.append({
            "q":      m.get("question", ""),
            "cid":    cid,
            "slug":   m.get("slug", ""),
            "yes":    yes_p,
            "no":     no_p,
            "liq":    liq,
            "vol24":  vol24,
            "end_dt": end_dt,
        })

print(f"  Загружено действующих рынков: {len(all_markets)}")

# ════════════════════════════════════════════════════════════
# 2. ФИЛЬТР СПОРТИВНЫХ РЫНКОВ
# ════════════════════════════════════════════════════════════

SPORT_KW = [
    "win", "beat", " vs ", "o/u", "over", "under", "total",
    "nba", "nhl", "nfl", "ufc", "mlb", "epl", "ucl",
    "fc ", " fc", "afc", "sc ",
    "tolima", "tijuana", "puebla", "tigres", "mazatlan",
    "atletico", "pachuca", "monterrey", "america", "chivas",
    "auxerre", "strasbourg", "lens", "metz",
    "gracheva", "tagger", "tennis", "match",
    "hurricanes", "stanley cup",
    "burnley", "arsenal", "chelsea", "liverpool", "tottenham",
    "manchester", "villa", "city", "united",
    "lakers", "celtics", "warriors", "nuggets", "bucks",
    "fight", "bout",
]

sport_markets = [
    m for m in all_markets
    if any(kw in m["q"].lower() for kw in SPORT_KW)
]
sport_markets.sort(key=lambda x: -x["liq"])

print(f"  Спортивных рынков: {len(sport_markets)}")

# ════════════════════════════════════════════════════════════
# 3. KELLY — РАСЧЁТ РАЗМЕРА СТАВКИ
# ════════════════════════════════════════════════════════════

def kelly_bet_for_side(
    side_price: float,
    p_win_side: float,
    bankroll: float = BANKROLL,
    frac: float = KELLY_FRAC,
    max_usd: float = MAX_BET_USD,
    min_usd: float = MIN_BET_USD,
) -> float:
    """
    Дробный Kelly для покупки доли по цене side_price при оценке вероятности выигрыша p_win_side.
    Одинаковая логика для YES и NO (цена = цена купленной доли).
    """
    p = min(max(p_win_side, 0.01), 0.95)
    c = side_price
    if c <= 0 or c >= 1:
        return 0.0
    b = (1 / c) - 1
    if b <= 0:
        return 0.0
    f = (p * b - (1 - p)) / b
    f = max(f, 0)
    bet = bankroll * f * frac
    bet = min(bet, max_usd)
    bet = max(bet, min_usd) if f > 0 else 0
    return round(bet, 2)

# Выбираем топ-5 рынков с Kelly > 0, разных категорий
bets_to_place: list[dict] = []
seen_ends: set[str] = set()

for m in sport_markets:
    if len(bets_to_place) >= 5:
        break
    # Направление: YES если цена < 0.60, NO если > 0.60; Kelly отдельно по стороне
    if m["yes"] < 0.60:
        side = "YES"
        bet_price = m["yes"]
        p_est = min(m["yes"] + EDGE_EST, 0.95)
    else:
        side = "NO"
        bet_price = m["no"]
        p_est = min((1.0 - m["yes"]) + EDGE_EST, 0.95)

    bet_usd = kelly_bet_for_side(bet_price, p_est)
    if bet_usd == 0:
        continue
    # Не дублируем матчи с одинаковой датой-временем
    end_key = m["end_dt"].strftime("%Y-%m-%d-%H") if m["end_dt"] else "no-date"
    # Разрешаем не более 2 ставок в одно окно
    if sum(1 for k in seen_ends if k == end_key) >= 2:
        continue
    seen_ends.add(end_key)

    win_usd = round(bet_usd * (1 / bet_price - 1), 2)
    p_true = p_est
    ev = round(p_true * win_usd - (1 - p_true) * bet_usd, 2)
    decimal_odds = round(1 / bet_price, 3)

    check_dt = (m["end_dt"] + timedelta(hours=2)) if m["end_dt"] else None

    bets_to_place.append({
        "q":           m["q"],
        "cid":         m["cid"],
        "slug":        m["slug"],
        "yes":         m["yes"],
        "no":          m["no"],
        "side":        side,
        "price":       bet_price,
        "odds":        decimal_odds,
        "bet_usd":     bet_usd,
        "win_usd":     win_usd,
        "ev":          ev,
        "liq":         m["liq"],
        "vol24":       m["vol24"],
        "end_dt":      m["end_dt"],
        "check_dt":    check_dt,
    })

total_risk = sum(b["bet_usd"] for b in bets_to_place)
total_ev   = sum(b["ev"]      for b in bets_to_place)

print(f"\n  Ставок отобрано: {len(bets_to_place)}")
print(f"  Суммарный риск: ${total_risk:.2f}")
print(f"  Суммарный EV: ${total_ev:+.2f}")

# ════════════════════════════════════════════════════════════
# 4. ЗАПИСЬ В EXCEL — ВКЛАДКА СТАВКИ
# ════════════════════════════════════════════════════════════

wb = load_workbook(EXCEL_PATH)
ws_bets  = wb["Ставки"]
ws_stats = wb["Статистика"]

HEADERS = [
    "№", "Condition ID", "Рынок (вопрос)",
    "Направление", "Коэф. (dec)",
    "Ставка $", "Цена доли", "Потенц. выигрыш $", "EV $",
    "Результат", "P&L $", "Статус",
    "Дата ставки", "Дата окончания", "Проверить в",
]

# Проверяем/обновляем заголовки
for c, h in enumerate(HEADERS, 1):
    ws_bets.cell(row=1, column=c, value=h)
hdr(ws_bets, 1, len(HEADERS))
ws_bets.row_dimensions[1].height = 32

# Находим первую пустую строку
next_row = 2
while ws_bets.cell(row=next_row, column=3).value not in (None, ""):
    next_row += 1

start_row = next_row

for i, b in enumerate(bets_to_place, 1):
    r = next_row
    next_row += 1
    # Зебра
    base = ZEBRA_FILL if i % 2 == 0 else PatternFill()

    bet_num  = start_row - 1 + i   # порядковый номер ставки
    side_fill = BLUE_FILL if b["side"] == "YES" else ORANGE_FILL

    end_str   = b["end_dt"].strftime("%Y-%m-%d %H:%M")  if b["end_dt"]   else ""
    check_str = b["check_dt"].strftime("%Y-%m-%d %H:%M") if b["check_dt"] else ""

    ws_bets.cell(row=r, column=1,  value=bet_num).alignment = CENTER
    ws_bets.cell(row=r, column=2,  value=b["cid"][:30])
    ws_bets.cell(row=r, column=3,  value=b["q"]).alignment = WRAP
    ws_bets.cell(row=r, column=4,  value=b["side"]).alignment = CENTER
    ws_bets.cell(row=r, column=5,  value=b["odds"]).number_format = "0.000"
    ws_bets.cell(row=r, column=6,  value=b["bet_usd"]).number_format = '#,##0.00'
    ws_bets.cell(row=r, column=7,  value=b["price"]).number_format = "0.000"
    ws_bets.cell(row=r, column=8,  value=b["win_usd"]).number_format = '#,##0.00'
    ws_bets.cell(row=r, column=9,  value=b["ev"]).number_format = '#,##0.00'
    ws_bets.cell(row=r, column=10, value="Ожидание").alignment = CENTER
    ws_bets.cell(row=r, column=11, value="").number_format = '#,##0.00'
    ws_bets.cell(row=r, column=12, value="Активна").alignment = CENTER
    ws_bets.cell(row=r, column=13, value=NOW.strftime("%Y-%m-%d %H:%M"))
    ws_bets.cell(row=r, column=14, value=end_str)
    ws_bets.cell(row=r, column=15, value=check_str)

    ws_bets.cell(row=r, column=4).fill  = side_fill
    ws_bets.cell(row=r, column=10).fill = YELLOW_FILL
    ws_bets.cell(row=r, column=12).fill = BLUE_FILL

    for c in range(1, len(HEADERS)+1):
        cell = ws_bets.cell(row=r, column=c)
        cell.border = THIN_BORDER
        if c not in (4, 10, 12):
            cell.fill = base
    ws_bets.row_dimensions[r].height = 32

# Ширина столбцов
col_w = {"A":4,"B":18,"C":52,"D":12,"E":10,"F":10,"G":9,"H":18,"I":9,
          "J":14,"K":10,"L":10,"M":18,"N":18,"O":18}
for col, w in col_w.items():
    ws_bets.column_dimensions[col].width = w

# ── Статистика: обновляем строку итогов ──────────────────────
ws_stats.cell(row=7, column=1, value="Дата последнего обновл.")
ws_stats.cell(row=7, column=2, value=NOW.strftime("%Y-%m-%d %H:%M UTC"))
ws_stats.cell(row=8, column=1, value="Активных ставок")
ws_stats.cell(row=8, column=2, value=len(bets_to_place))
ws_stats.cell(row=9, column=1, value="Суммарный риск $")
ws_stats.cell(row=9, column=2, value=round(total_risk, 2))
ws_stats.cell(row=10, column=1, value="Суммарный EV $")
ws_stats.cell(row=10, column=2, value=round(total_ev, 2))
for r in range(7, 11):
    for c in [1, 2]:
        ws_stats.cell(row=r, column=c).border = THIN_BORDER
        ws_stats.cell(row=r, column=c).font = Font(bold=(c==1))

# Журнал
log_row = 12
while ws_stats.cell(row=log_row, column=1).value not in (None, ""):
    log_row += 1
    if log_row > 300: break
ws_stats.cell(row=log_row, column=1, value=NOW.strftime("%Y-%m-%d %H:%M"))
ws_stats.cell(row=log_row, column=2,
    value=f"Добавлено {len(bets_to_place)} ставок. "
          f"Риск=${total_risk:.2f}, EV=${total_ev:+.2f}. "
          f"Рынки: {', '.join(b['q'][:30] for b in bets_to_place)}")
ws_stats.cell(row=log_row, column=2).alignment = WRAP
for c in [1, 2]:
    ws_stats.cell(row=log_row, column=c).border = THIN_BORDER
ws_stats.cell(row=log_row, column=1).fill = BLUE_FILL

wb.save(EXCEL_PATH)
print(f"\n  Excel сохранён: {EXCEL_PATH}\n")

# ════════════════════════════════════════════════════════════
# 5. ВЫВОД ОТЧЁТА В КОНСОЛЬ
# ════════════════════════════════════════════════════════════

print("=" * 82)
print(f"  ПОСТАВЛЕННЫЕ СТАВКИ — {NOW.strftime('%Y-%m-%d %H:%M UTC')}")
print("=" * 82)
for i, b in enumerate(bets_to_place, 1):
    end_str = b["end_dt"].strftime("%d.%m.%Y %H:%M UTC") if b["end_dt"] else "?"
    print(f"\n  #{i}  [{b['side']}]  Коэф. {b['odds']:.3f}")
    print(f"  {b['q'][:75]}")
    print(f"  Ставка: ${b['bet_usd']:.2f}   "
          f"Выигрыш при победе: +${b['win_usd']:.2f}   "
          f"EV: ${b['ev']:+.2f}")
    print(f"  Цена доли: {b['price']:.3f}   "
          f"Ликвидность: ${b['liq']:,.0f}   "
          f"Окончание: {end_str}")
    chk = b["check_dt"].strftime("%d.%m.%Y %H:%M") if b["check_dt"] else "?"
    print(f"  Проверить результат: {chk}")

print()
print("-" * 82)
print(f"  Суммарный риск:  ${total_risk:.2f}")
print(f"  Суммарный EV:    ${total_ev:+.2f}")
print(f"  Bankroll:        ${BANKROLL:.2f}")
print(f"  % bankroll под риском: {total_risk/BANKROLL*100:.1f}%")
print("-" * 82)
print()
print("  НАПОМИНАНИЕ: ставки зарегистрированы как БУМАЖНЫЕ.")
print("  Для реального размещения — настрой POLYMARKET_PRIVATE_KEY в .env")
print("  и запусти: python main.py --mode auto --bankroll 1000")
print()
